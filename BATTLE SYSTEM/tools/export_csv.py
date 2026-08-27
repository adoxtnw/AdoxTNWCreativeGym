#!/usr/bin/env python3
"""
Export every sheet of the config workbook to CSV.

    python3 tools/export_csv.py [out-folder]

The spreadsheet is the source of truth, but build_data.py reads CSV (the format
Google Sheets exports). This closes the loop locally, so the whole chain is:

    build_workbook.py  ->  battle-system-config.xlsx  ->  export_csv.py  ->  build_data.py

Files are named the way Google names them ("battle-system-config - rules.csv") so
build_data.py matches them without special-casing.
"""
import sys, os, csv
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BOOK = os.path.join(HERE, "..", "config", "battle-system-config.xlsx")

def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "..", "config", "csv")
    os.makedirs(out, exist_ok=True)
    wb = openpyxl.load_workbook(BOOK, data_only=True)
    n = 0
    for name in wb.sheetnames:
        ws = wb[name]
        path = os.path.join(out, f"battle-system-config - {name}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                if all(c is None for c in row):
                    continue
                w.writerow(["" if c is None else c for c in row])
        n += 1
    print(f"exported {n} sheets -> {out}")

if __name__ == "__main__":
    main()
