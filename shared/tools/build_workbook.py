#!/usr/bin/env python3
"""
Rebuild config/battle-system-config.xlsx from scratch.

    python3 tools/build_workbook.py

This is the generator for the WORKBOOK ITSELF — sheet layout, column sets,
LIVE/planned markers, validation formulas, and the seed content. Run it when
you want to add a sheet or a column. Day-to-day balancing happens in Google
Sheets instead; this would overwrite that.

Dialogue rows are imported from sources/neuro_metro_avui_enemy_dialogues.csv.
"""
import os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from _paths import app_root, book_path, csv_dir
APP  = app_root()
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WB = openpyxl.Workbook()
WB.remove(WB.active)

F = "Arial"
HDR_FILL   = PatternFill("solid", fgColor="1E1E26")
HDR_FONT   = Font(name=F, bold=True, color="F0ECE0", size=10)
LIVE_FILL  = PatternFill("solid", fgColor="D9EAD3")   # column read by the prototype today
PLAN_FILL  = PatternFill("solid", fgColor="FFF2CC")   # reserved for future design
NOTE_FONT  = Font(name=F, italic=True, color="666666", size=9)
BODY       = Font(name=F, size=10)
BOLD       = Font(name=F, bold=True, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def sheet(name, headers, live, rows, widths=None, notes=None):
    """headers: list of column names. live: set of names the game reads today."""
    ws = WB.create_sheet(name)
    if notes:
        ws["A1"] = notes; ws["A1"].font = NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers),6))
        hrow = 3
    else:
        hrow = 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORD
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        tag = ws.cell(row=hrow+1, column=c, value="LIVE" if h in live else "planned")
        tag.font = Font(name=F, size=8, bold=(h in live),
                        color="274E13" if h in live else "7F6000")
        tag.fill = LIVE_FILL if h in live else PLAN_FILL
        tag.border = BORD
        tag.alignment = Alignment(horizontal="left")
    for r, row in enumerate(rows, hrow+2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.font = BODY; cell.border = BORD
    for c, h in enumerate(headers, 1):
        w = (widths or {}).get(h, max(11, min(34, len(h) + 5)))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=hrow+2, column=2)
    return ws, hrow

# ───────────────────────────── README ─────────────────────────────
rd = WB.create_sheet("README")
lines = [
 ("NEURO-METRO: AVUI — GAME CONFIG", 16, True, "F0ECE0", "1E1E26"),
 ("", 10, False, None, None),
 ("This workbook is the source of truth for all game content. The prototype reads no", 10, False, None, None),
 ("content values of its own — every number, ability, unit and rule comes from here.", 10, False, None, None),
 ("", 10, False, None, None),
 ("HOW THE ROUND TRIP WORKS", 12, True, None, None),
 ("1.  Edit any sheet in Google Sheets.", 10, False, None, None),
 ("2.  File > Download > Comma-separated values, one CSV per sheet.", 10, False, None, None),
 ("3.  Put the CSVs in a folder and run:  python3 tools/build_data.py <folder>", 10, False, None, None),
 ("4.  That regenerates data.js. Reload the prototype. Nothing else to touch.", 10, False, None, None),
 ("", 10, False, None, None),
 ("COLUMN STATUS — the row under each header", 12, True, None, None),
 ("LIVE      (green)   the prototype reads this column today. Renaming it breaks the game.", 10, False, None, None),
 ("planned   (yellow)  reserved for design you have not built yet. Inert until wired up;", 10, False, None, None),
 ("                    safe to fill in early. Tell me when you want one activated.", 10, False, None, None),
 ("", 10, False, None, None),
 ("CONVENTIONS", 12, True, None, None),
 ("ids          UPPER_SNAKE. Referenced by other sheets, so renaming one means updating", 10, False, None, None),
 ("             every sheet that points at it. The 'checks' sheet catches broken references.", 10, False, None, None),
 ("lists        pipe-separated in a single cell:  ATK_ANGER|ATK_JOY|DEFEND", 10, False, None, None),
 ("booleans     1 or 0  (TRUE/FALSE also accepted)", 10, False, None, None),
 ("blank        means 'none' / 'typeless'. An ability with no emotion never matches a layer.", 10, False, None, None),
 ("*            wildcard, in the matchups sheet only. Matches any emotion.", 10, False, None, None),
 ("enabled      set to 0 to switch a row off without deleting it.", 10, False, None, None),
 ("", 10, False, None, None),
 ("ADDING THINGS — no code needed", 12, True, None, None),
 ("a new ability      add a row to 'abilities', then add its id to a unit's pool", 10, False, None, None),
 ("a new enemy        add a row to 'units'", 10, False, None, None),
 ("a new emotion      add a row to 'emotions' with a hex colour, then matchup rows for it", 10, False, None, None),
 ("a new synergy      add a row to 'matchups' — this is the main lever for interactions", 10, False, None, None),
 ("retune anything    'rules'", 10, False, None, None),
 ("", 10, False, None, None),
 ("SHEETS", 12, True, None, None),
 ("emotions         the six emotional types and their palette colours", 10, False, None, None),
 ("abilities        everything a unit can do", 10, False, None, None),
 ("matchups         attack emotion x layer emotion -> damage and charge. Your synergy table.", 10, False, None, None),
 ("units            player and enemies: stats, starting layers, ability pool", 10, False, None, None),
 ("layer_types      per-layer properties. Reserved — layers are plain emotions today.", 10, False, None, None),
 ("status_effects   buffs and debuffs. Reserved.", 10, False, None, None),
 ("synergies        line-level combos (sequences, counts, positions). Reserved.", 10, False, None, None),
 ("dialogue         what each enemy says, per persona and battle state", 10, False, None, None),
 ("rules            global tunables", 10, False, None, None),
 ("sounds           runtime-synthesised SFX. No audio files.", 10, False, None, None),
 ("checks           live validation. Every row should read OK.", 10, False, None, None),
]
for i,(txt,size,bold,fg,bg) in enumerate(lines, 1):
    c = rd.cell(row=i, column=1, value=txt)
    c.font = Font(name=F, size=size, bold=bold, color=fg or "000000")
    if bg: c.fill = PatternFill("solid", fgColor=bg)
rd.column_dimensions["A"].width = 100

# ───────────────────────────── emotions ─────────────────────────────
EMO = [
 # OFFICIAL palette, straight from Bars_Reference_001.svg. This list's ORDER is
 # also the gradient order (the reference puts them at even 20% stops).
 # the last field is the glyph tiled behind this emotion's Loadout button
 ("ANGER",   "Anger",   "anger",   "#e53859","#2a0810","Hot, direct, aggressive.","BOLT"),
 ("SURPRISE","Surprise","surprise","#724082","#170a1b","Sudden, disruptive, disorienting.","BURST"),
 ("DISGUST", "Disgust", "disgust", "#56a36a","#0c1f13","Rejecting, corrosive, contemptuous.","ROT"),
 ("JOY",     "Joy",     "joy",     "#fcc336","#2a1f06","Bright, volatile, energising.","SPARK"),
 ("SADNESS", "Sadness", "sadness", "#3d66c1","#0a1229","Heavy, draining, persistent.","DROP"),
 ("FEAR",    "Fear",    "fear",    "#929fa5","#1a1e20","Cold, freezing, evasive.","EYE"),
]
sheet("emotions",
  ["id","name","token","hex","bg_hex","short","icon","sfx","description","enabled","notes"],
  {"id","name","token","hex","bg_hex","icon","sfx"},
  [dict(id=i,name=n,token=t,hex=h,bg_hex=b,short=i[:3],icon=ic,sfx="hit_"+i.lower(),
        description=d,enabled=1,notes="")
   for i,n,t,h,b,d,ic in EMO],
  widths={"description":40,"notes":30},
  notes=("One row per emotional type. 'token' is the CSS palette variable name; 'hex' drives every colour "
         "in the UI for this emotion; 'bg_hex' tints the battle backdrop when an enemy of this type is fought."))

# ───────────────────────────── abilities ─────────────────────────────
ABIL_COLS = ["id","name","emotion","cost","kind","power","charge","hits_layer","icon",
             "target","reach","self_ms","self_ec","ec_push_target","ec_drain_target",
             "heal","shield_gain","pierce_shield","ignore_layer","repeat","cooldown",
             "uses","blurb","action","emotions","crit_chance","status_apply","status_chance","status_duration",
             "combo_tag","requires_prev_tag","synergy_group","wild_target",
             "rarity","unlock","enabled","suggested_cost","cost_delta","notes"]
ABIL_LIVE = {"id","name","emotion","cost","kind","power","charge","hits_layer","icon","self_ms","cooldown","uses","blurb","action","emotions","crit_chance","status_apply","status_duration"}
def ab(**k):
    row = {c:"" for c in ABIL_COLS}
    row.update(dict(target="ENEMY", reach="SINGLE", charge=0, repeat=1, cooldown=1,
                    uses=3, action=0, emotions="", crit_chance="", status_chance=0, status_duration=0,
                    pierce_shield=0, ignore_layer=0, enabled=1,
                    self_ms=0, self_ec=0, ec_push_target=0, ec_drain_target=0,
                    heal=0, shield_gain=0, wild_target="LOWEST_MS", rarity="COMMON"))
    row.update(k); return row
ABILS = [
 ab(id="ATK_ANGER", blurb='Attack. Cracks the outermost {LAYER}.',  name="Heated Punch",  emotion="ANGER",   cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="BOLT"),
 ab(id="ATK_SADNESS", blurb='Attack. Cracks the outermost {LAYER}.',name="Cold Shoulder",emotion="SADNESS", cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="DROP"),
 ab(id="ATK_JOY", blurb='Attack. Cracks the outermost {LAYER}.',    name="Manic Grin",    emotion="JOY",     cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="SPARK"),
 ab(action=1, id="DEFEND", blurb='Defence. Blocks the next hit outright.',     name="Block", emotion="",        cost=10, kind="SHIELD", power=1,  hits_layer=0, icon="SHIELD",
    target="SELF", wild_target="SELF", shield_gain=1),
 ab(uses=2, id="HVY_ANGER", blurb='Attack. *Charges* first, then hits for triple.',  name="Rage",   emotion="ANGER",   cost=45, kind="DAMAGE", power=90, charge=2, hits_layer=1, icon="BOLT"),
 ab(uses=2, id="HVY_SADNESS", blurb='Attack. *Charges* first, then hits for triple.',name="Grief",  emotion="SADNESS", cost=45, kind="DAMAGE", power=90, charge=2, hits_layer=1, icon="DROP"),
 ab(uses=2, id="HVY_JOY", blurb='Attack. The hardest hit there is, after a long *charge*.',    name="Mania",  emotion="JOY",     cost=60, kind="DAMAGE", power=130,charge=3, hits_layer=1, icon="SPARK"),
 ab(action=1, id="RECHARGE", blurb='Buff. Trades your own {MS} for {EC}.',   name="Recharge",emotion="",       cost=0,  kind="CHARGE", power=20, hits_layer=0, icon="CHARGE",
    target="SELF", wild_target="SELF", self_ms=5,
    notes="Gains 20 EC and costs 5 of your own MS — which also lowers your ceiling."),
 ab(id="GEN_DISGUST", blurb='Defence. Grows a {LAYER} that *never regrows* once broken.', name="Bile",   emotion="DISGUST", cost=15, kind="ADDLAYER", power=1, hits_layer=0, icon="ROT",
    target="SELF", wild_target="SELF", uses=2,
    notes="Grows one extra Disgust layer. Grown layers are TEMPORARY: once broken they never regrow."),
 ab(id="GEN_ANGER", blurb='Defence. Grows a {LAYER} that *never regrows* once broken.',   name="Bristle",emotion="ANGER",   cost=15, kind="ADDLAYER", power=1, hits_layer=0, icon="ROT",
    target="SELF", wild_target="SELF", uses=2,
    notes="Grows one extra Anger layer. Grown layers are TEMPORARY: once broken they never regrow."),
 # Joy had exactly two abilities, both of them attacks, which left LO_JOY half
 # empty and made "mostly its own type" impossible for a weak Joy enemy without
 # handing it the hardest hit in the game. Same shape as Bristle and Bile.
 ab(id="GEN_JOY", blurb='Defence. Grows a {LAYER} that *never regrows* once broken.', name="Good Vibes", emotion="JOY", cost=15, kind="ADDLAYER", power=1, hits_layer=0, icon="SPARK",
    target="SELF", wild_target="SELF", uses=2,
    notes="Grows one extra Joy layer. Grown layers are TEMPORARY: once broken they never regrow."),
 ab(id="ROT", blurb='Debuff. Target cannot regrow *2* {LAYERS} for *2 turns*.',         name="Fester",    emotion="DISGUST", cost=25, kind="DEBUFF",   power=0, hits_layer=0, icon="ROT",
    uses=2, status_apply="NO_REGEN", status_duration=2,
    notes="For 2 turns the target cannot regrow 2 of its broken layers."),
 ab(id="INFLICT_SAD", blurb='Status. Target attacks *itself* each round for *2 turns*.', name="Self-Harm", emotion="SADNESS", cost=30, kind="DEBUFF", power=0, hits_layer=0, icon="DROP",
    uses=2, status_apply="SAD", status_duration=2,
    notes="Each round end the target turns one of its own attacks on itself. NOTE: distinct from "
          "SELF_HARM, which is the Overload-forced station."),
 ab(id="BLIND", blurb='Debuff. Target *misses half* its attacks for *2 turns*.',       name="Blinded by Hate", emotion="ANGER", cost=30, kind="DEBUFF", power=0, hits_layer=0, icon="EYE",
    uses=2, status_apply="BLINDED", status_duration=2,
    notes="ENEMY ability. The target misses half its attacks for 2 turns."),
 # SURPRISE had a Loadout and a metro line and no abilities at all, so a Surprise
 # enemy could only fight with borrowed emotions. Same shape as the Anger and
 # Sadness kits — a basic, a charged heavy and a debuff — so nothing about the
 # costing or the shot counts is new; only the emotion is.
 ab(id="ATK_SURPRISE", blurb='Attack. Cracks the outermost {LAYER}.', name="Sucker Punch", emotion="SURPRISE", cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="BURST"),
 ab(uses=2, id="HVY_SURPRISE", blurb='Attack. *Charges* first, then hits for triple.', name="Whiplash", emotion="SURPRISE", cost=45, kind="DAMAGE", power=90, charge=2, hits_layer=1, icon="BURST"),
 ab(id="STARTLE", blurb='Debuff. Target *misses a third* of its attacks for *2 turns*.', name="Out of Nowhere", emotion="SURPRISE", cost=25, kind="DEBUFF", power=0, hits_layer=0, icon="BURST",
    uses=2, status_apply="RATTLED", status_duration=2,
    notes="Surprise's debuff. Softer than BLIND and cheaper, because Surprise pays for it "
          "with a weaker basic economy rather than with a bigger bill."),
 ab(cooldown=0, uses=0, id="SELF_HARM", blurb='Status. Forced by {OVERLOAD}; costs you {MS}.',  name="Self Harm",emotion="",      cost=0,  kind="SELFHARM", power=25, hits_layer=0, icon="WARN",
    target="SELF", wild_target="SELF", rarity="OVERLOAD", enabled=1,
    notes="OVERLOAD ONLY. Forced into your line when Charge passes your ceiling. Cannot be moved or removed."),
 ab(cooldown=0, uses=0, id="FEED", blurb='Status. Forced by {OVERLOAD}; *heals* your opponent.',       name="Feed",    emotion="",       cost=0,  kind="FEED",     power=30, hits_layer=0, icon="DROP",
    target="ENEMY", wild_target="AS_WRITTEN", rarity="OVERLOAD", enabled=1,
    notes="OVERLOAD ONLY. Heals your opponent. Cannot be moved or removed."),
]
ws_ab, hr = sheet("abilities", ABIL_COLS, ABIL_LIVE, ABILS,
  widths={"notes":34,"combo_tag":16,"requires_prev_tag":18,"synergy_group":16,"status_apply":16},
  notes=("One row per ability. kind: DAMAGE | SHIELD (add more kinds as you design them). "
         "power = damage for DAMAGE, shield charges for SHIELD. Leave 'emotion' blank for typeless — "
         "a typeless ability never matches a layer. 'suggested_cost' is a formula: a costing sanity check, not a rule."))
first = hr+2
for r in range(first, first+len(ABILS)):
    ws_ab.cell(row=r, column=ABIL_COLS.index("suggested_cost")+1).value = (
        f"=ROUND(G{r}*0.6 + N{r}*0.4 + P{r}*0.7 + Q{r}*8 + ABS(L{r})*0.3, 0)")
    ws_ab.cell(row=r, column=ABIL_COLS.index("cost_delta")+1).value = (
        f"=IFERROR(D{r}-AG{r},\"\")")
    for cc in ("suggested_cost","cost_delta"):
        ws_ab.cell(row=r, column=ABIL_COLS.index(cc)+1).font = BODY

# ───────────────────────────── matchups ─────────────────────────────
MU_COLS = ["attack_emotion","layer_emotion","priority","dmg_mult","ec_mult","label",
           "tag_class","sound","rotates_layer","self_ec","status_apply","enabled","notes"]
MU_LIVE = {"attack_emotion","layer_emotion","priority","dmg_mult","ec_mult","label","tag_class"}
def mu(a,l,p,d,e,lab,cls,snd,note="",rot="",sec=0,st=""):
    return dict(attack_emotion=a, layer_emotion=l, priority=p, dmg_mult=d, ec_mult=e,
                label=lab, tag_class=cls, sound=snd, rotates_layer=rot, self_ec=sec,
                status_apply=st, enabled=1, notes=note)
MUS = [
 mu("*","*",0,1,0,"OFF TYPE","off","hit","Fallback. Every pair you do not name lands here."),
 mu("*","NONE",5,2,0,"EXPOSED!","dmg","hit","No layers left: double damage, no charge."),
]
for e,*_ in EMO:
    MUS.append(mu(e,e,10,0.5,1,"ABSORBED!","absorb","absorb","Same emotion: the layer drinks it."))
sheet("matchups", MU_COLS, MU_LIVE, MUS,
  widths={"notes":52,"label":16},
  notes=("THE SYNERGY TABLE — the main lever for designing interactions. One row per "
         "attack-emotion x layer-emotion pair. '*' is a wildcard; 'NONE' in layer_emotion means the "
         "target has no layers. When several rows match, the HIGHEST priority wins, so specific rows "
         "beat wildcards. Add a row like  JOY | SADNESS | 20 | 2.0 | 0 | CATHARSIS  and that pairing "
         "immediately behaves differently — no code change."))

# ───────────────────────────── units ─────────────────────────────
U_COLS = ["id","name","emotion","tier","max_ms","start_ec_pct","layers","pool","line_dir",
          "line_cap","max_bonus_slots","loadouts","spawn_lines","drops",
          "ai_profile","init","start_shield","max_layers_override","tags","enabled","notes"]
U_LIVE = {"id","name","emotion","tier","max_ms","start_ec_pct","layers","pool","line_dir","line_cap",
          "max_bonus_slots","loadouts","spawn_lines","drops"}
def en(**k):
    """An enemy row. Everything an enemy shares with every other enemy is here, so a
    new one is the handful of cells that actually make it that enemy."""
    row = dict(tier="REGULAR", start_ec_pct=0.40, line_dir=-1, line_cap=3, max_bonus_slots=6,
               loadouts="", ai_profile="GREEDY_MAX_DAMAGE", init=8, start_shield=0,
               max_layers_override="", tags="ENEMY", enabled=1, notes="")
    row.update(k); return row
UNITS_ROWS = [
 dict(id="player", name="You", emotion="", tier="", spawn_lines="", max_ms=400, start_ec_pct=0.50,
      layers="JOY|SADNESS", pool="ATK_ANGER|ATK_SADNESS|ATK_JOY|DEFEND|RECHARGE|HVY_ANGER|HVY_SADNESS|HVY_JOY|GEN_DISGUST|GEN_ANGER|ROT|INFLICT_SAD", line_dir=1,
      line_cap=3, max_bonus_slots=6, loadouts="LO_ANGER|LO_SADNESS|LO_JOY",
      drops="",
      ai_profile="", init=10, start_shield=0, max_layers_override="", tags="PLAYER", enabled=1, notes=""),
 en(id="enemy", name="The Commuter", emotion="ANGER", tier="REGULAR", max_ms=250,
    layers="ANGER|ANGER|SADNESS",
    pool="ATK_ANGER|ATK_SADNESS|ATK_JOY|DEFEND|RECHARGE|HVY_ANGER|HVY_SADNESS|HVY_JOY|BLIND",
    spawn_lines="L1:1.0|*:0.6",
    drops="CRYSTAL:2:0.75|SEGMENT:3:0.55|ORB:1:0.40",
    notes="AI reads the matchups sheet, so retuning it retunes the AI. `drops` is what "
          "beating this one may leave. The `*` in spawn_lines is what stops L3, L4 and L6 "
          "being empty of enemies until they have units of their own — the Commuter rides "
          "every line, which is the joke and also the fallback."),
 en(id="enemy_anger_strong", name="The Enforcer", emotion="ANGER", tier="STRONG",
    max_ms=330, start_ec_pct=0.50, line_cap=3,
    layers="ANGER|ANGER|ANGER|SADNESS",
    pool="ATK_ANGER|HVY_ANGER|GEN_ANGER|BLIND|ATK_SADNESS|DEFEND|RECHARGE",
    spawn_lines="L1:0.5",
    drops="CRYSTAL:3:0.85|SEGMENT:4:0.70|ORB:1:0.55",
    notes="Anger, turned up: one more layer than the Commuter and a third more stamina, and "
          "its pool is nearly all Anger, so an Anger-layered player is drinking most of it "
          "while anyone else is not."),
 en(id="enemy_surprise", name="The Interruption", emotion="SURPRISE", tier="REGULAR",
    max_ms=240,
    layers="SURPRISE|SURPRISE|ANGER",
    pool="ATK_SURPRISE|HVY_SURPRISE|STARTLE|ATK_ANGER|DEFEND|RECHARGE",
    spawn_lines="L2:1.0",
    drops="CRYSTAL:2:0.75|SEGMENT:3:0.55|ORB:1:0.40",
    notes="L2's own. Slightly under the Commuter on stamina because STARTLE is worth more "
          "than it costs when it lands."),
 en(id="enemy_surprise_strong", name="The Reversal", emotion="SURPRISE", tier="STRONG",
    max_ms=320, start_ec_pct=0.50,
    layers="SURPRISE|SURPRISE|SURPRISE|JOY",
    pool="ATK_SURPRISE|HVY_SURPRISE|STARTLE|GEN_ANGER|ATK_JOY|DEFEND|RECHARGE",
    spawn_lines="L2:0.45",
    drops="CRYSTAL:3:0.85|SEGMENT:4:0.70|ORB:1:0.55",
    notes="Three Surprise layers deep. GEN_ANGER is in the pool so it can grow a layer that "
          "does NOT absorb Surprise, which is the counter to a player who came dressed for it."),
 en(id="enemy_sadness_weak", name="The Straggler", emotion="SADNESS", tier="WEAK",
    max_ms=150, start_ec_pct=0.30, line_cap=2, max_bonus_slots=4,
    layers="SADNESS|SADNESS",
    pool="ATK_SADNESS|INFLICT_SAD|DEFEND|RECHARGE",
    spawn_lines="L2:0.15|L5:0.15",
    drops="CRYSTAL:1:0.45|SEGMENT:2:0.40",
    notes="No heavy in the pool, on purpose: a WEAK enemy chips. Two slots, two layers, and "
          "INFLICT_SAD is the one thing it can do that you will remember. RARE EVEN AT HOME "
          "— the sheet says L5 at 0.15, not 1.0, because it was asked for as an uncommon "
          "sight on both its lines. Raise the L5 cell to make it Line 5's regular."),
 en(id="enemy_joy_weak", name="The Reveller", emotion="JOY", tier="WEAK",
    max_ms=150, start_ec_pct=0.30, line_cap=2, max_bonus_slots=4,
    layers="JOY|JOY",
    pool="ATK_JOY|GEN_JOY|DEFEND|RECHARGE",
    spawn_lines="L2:0.15",
    drops="CRYSTAL:1:0.45|SEGMENT:2:0.40",
    notes="L2 ONLY, and rarely — it is not on L4, its own colour's line, because that is how "
          "it was asked for. Reads as Line 2 being where the wrong people end up. Add "
          "`|L4:1.0` to give Joy its own regular."),
]
sheet("units", U_COLS, U_LIVE, UNITS_ROWS,
  widths={"layers":26,"pool":40,"spawn_lines":20,"notes":42},
  notes=("One row per combatant. 'layers' is outermost-first and rotates as it takes hits. "
         "'pool' is which abilities this unit may use. Add enemies by adding rows. "
         "`drops` is what BEATING this unit may leave, as kind:amount:chance - "
         "CRYSTAL:2:0.75|SEGMENT:3:0.55|ORB:1:0.40 means up to two crystals three quarters "
         "of the time, up to three Track Segments just over half, and a Stamina Orb (which "
         "gives back orbMsPct of MaxMS) two times in five. Each is rolled on its own, so an "
         "enemy can leave everything, something, or nothing. Blank means nothing ever. "
         "`tier` is WEAK / REGULAR / STRONG: it is not a stat, it is what the enemy LOOKS "
         "like - a weak one is concentric triangles in battle and a small dart on the ride, "
         "a strong one is a seven-pointed star and a large spined shape. "
         "`spawn_lines` is where the map may produce it, as line:weight - L2:0.15|L5:0.15 "
         "means an uncommon sight on either. `*` is every line, and is what keeps a line "
         "with no units of its own from being empty."))

# ───────────────────────────── layer_types (reserved) ─────────────────────────────
LT_COLS = ["id","emotion","display_name","durability","on_break_effect","on_break_value",
           "passive_effect","passive_value","regrow_delay_ms","tags","enabled","notes"]
sheet("layer_types", LT_COLS, set(),
  [dict(id="EXAMPLE_BRITTLE_ANGER", emotion="ANGER", display_name="Brittle Anger", durability=1,
        on_break_effect="EC_TO_OWNER", on_break_value=10, passive_effect="DMG_TAKEN_MULT",
        passive_value=1.25, regrow_delay_ms=380, tags="EXAMPLE", enabled=0,
        notes="EXAMPLE ROW — shows the intended format. enabled=0, so it is ignored.")],
  widths={"notes":54,"on_break_effect":20,"passive_effect":20},
  notes=("RESERVED. Layers are plain emotions today — units.layers holds emotion ids. This sheet exists "
         "for when a layer needs properties of its own (taking several hits to break, doing something "
         "when it breaks, or granting a passive while it is outermost)."))

# ───────────────────────────── status_effects (reserved) ─────────────────────────────
SE_COLS = ["id","name","duration","icon","color","blurb","block_regen","miss_chance","self_hits","crit_mult",
           "dmg_taken_mult","dmg_dealt_mult","ec_gain_mult","ms_per_turn","ec_per_turn",
           "blocks_actions","stacking","max_stacks","enabled","notes"]
SE_LIVE = {"id","name","duration","icon","color","blurb","block_regen","miss_chance","self_hits","crit_mult","enabled"}
def se(**k):
    row = {c:"" for c in SE_COLS}
    row.update(dict(duration=2, block_regen=0, miss_chance=0, self_hits=0, crit_mult=0,
                    dmg_taken_mult=1, dmg_dealt_mult=1, ec_gain_mult=1,
                    ms_per_turn=0, ec_per_turn=0, blocks_actions=0,
                    stacking="REFRESH", max_stacks=1, enabled=1))
    row.update(k); return row
SE_ROWS = [
 se(id="NO_REGEN", name="Rotting", duration=2, icon="ROT", color="#56a36a", block_regen=2,
    blurb="*2* of your broken {LAYERS} are being held down. They stay broken until this wears off.",
    notes="Holds this many broken layers down: they stay broken while it lasts."),
 se(id="SAD",      name="Sad",     duration=2, icon="DROP", color="#3d66c1", self_hits=1,
    blurb="At the end of every round one of your own attacks is turned *on yourself*.",
    notes="Each round end the victim turns one of its OWN attacks on itself."),
 se(id="OVERLOAD", name="Overload", duration=0, icon="WARN", color="#e53859",
    blurb="Your {EC} is over your {MS}. Your line is being invaded until you burn it down.",
    notes="DERIVED, not timed: it lasts while ec > ms, so it is never stored in unit.statuses."),
 se(id="BLINDED",  name="Blinded", duration=2, icon="EYE", color="#e53859", miss_chance=0.5,
    blurb="{ANGER} has blotted out your aim. *Half* of your attacks miss outright.",
    notes="The victim fluffs this fraction of its attacks."),
 se(id="RATTLED",  name="Rattled", duration=2, icon="BURST", color="#724082", miss_chance=0.35,
    blurb="{SURPRISE} came from nowhere. *A third* of your attacks go wide.",
    notes="Surprise's own miss status. missChance() takes the HIGHEST of everything on a "
          "unit rather than summing, so Rattled and Blinded together are just Blinded."),
]
sheet("status_effects", SE_COLS, SE_LIVE, SE_ROWS,
  widths={"notes":52},
  notes=("One row per status. Abilities apply these by id through status_apply / "
         "status_duration. Add a status by adding a row and teaching Kinds what reads it."))


# ───────────────────────────── synergies (reserved) ─────────────────────────────
SY_COLS = ["id","name","trigger","subject","operator","value","window","effect",
           "effect_value","applies_to","priority","enabled","notes"]
sheet("synergies", SY_COLS, set(),
  [dict(id="EXAMPLE_TRIPLE", name="Monochrome Line", trigger="LINE_CONTAINS", subject="EMOTION_COUNT",
        operator=">=", value=3, window="LINE", effect="DMG_MULT", effect_value=1.5,
        applies_to="ALL_STATIONS", priority=10, enabled=0,
        notes="EXAMPLE ROW — three stations of one emotion in a line hit 50% harder. enabled=0."),
   dict(id="EXAMPLE_CHAIN", name="Grief After Rage", trigger="PREV_STATION", subject="EMOTION",
        operator="==", value="ANGER", window="STATION", effect="EC_MULT", effect_value=2,
        applies_to="THIS_STATION", priority=20, enabled=0,
        notes="EXAMPLE ROW — a station firing straight after an Anger station gives double charge. enabled=0.")],
  widths={"notes":56},
  notes=("RESERVED. Line-level combos: things that depend on what else is in the line, or on what "
         "fired just before. trigger/subject/operator/value form the condition; effect/effect_value "
         "form the result. Tell me the ones you want and I will wire the vocabulary you use here."))

# ───────────────────────────── rules ─────────────────────────────
RULES = [
 ("lineCap",3,"FALLBACK line size, used only if a unit row leaves line_cap blank."),
 ("maxLayers",6,"Layer slots per unit."),
 ("restEc",15,"Charge gained by departing with an empty line."),
 ("ecBasis","POWER","POWER = charge figured from the ability's base power. DEALT = from damage actually dealt."),
 ("noLayerDmgMult",1,"Superseded by the matchups sheet; kept as a fallback."),
 ("noLayerEcMult",0,"Superseded by the matchups sheet; kept as a fallback."),
 ("shieldRotatesLayer",0,"Whether a blocked attack still rotates the layer queue."),
 ("lowMsWarnPct",0.18,"Below this fraction of max MS the ceiling chevron pulses red."),
 ("clashSeconds",4,"Length of the death sequence."),
 ("layerRegenAtTurnEnd",1,"Broken layers stay gone until the round ends, then regrow at the back of the queue."),
 ("chargeStepMs",260,"How long each charge segment holds before the next."),
 ("flyMs",120,"How long a station takes to fly across and strike the target."),
 ("typeMs",32,"Delay between letters of dialogue."),
 ("dialogueHoldMs",3000,"How long a finished line stays on screen before it fades."),
 ("lowHpTalkPct",0.2,"Fraction of max MS below which the winning / losing lines fire."),
 ("overloadSlotPer",0.25,"Every this-much overflow (as a fraction of max MS) forces one more corrupted slot."),
 ("aiVarietyChance",0.22,"Chance the AI takes a random affordable attack instead of its best one."),
 ("aiDebuffChance",0.5,"Chance the AI spends a slot on a debuff the target is not already suffering."),
 ("aiGrowChance",0.45,"How often the enemy takes an ADDLAYER ability when one is worth a slot. Like the debuff branch, this exists because the AI scores on damage per slot and a layer deals none - without it Bristle, Bile and Good Vibes sit in the pool and are never once chosen."),
 ("aiChargeBias",0.55,"Chance the AI prefers an ability with charge segments when it can afford one."),
 ("shuffleLayersEachRound",1,"Re-order every unit's layer queue at the start of each round."),
 ("maxExtraSlots",6,"Hard cap on extra slots of any kind a line may carry."),
 ("loadoutPatternAlpha",0.13,"How faint the emotion glyph tiled behind a Loadout button is."),
 ("loadoutSlots",4,"Ability slots in one Loadout, and so cells per page."),
 ("equippedSlots",3,"Loadouts the player carries into a battle."),
("slotArriveMs",260,"Gap between one temporary slot flying in and the next."),
 ("impactFlashMs",45,"White frame on a hit."),
 ("impactShakeMs",85,"Shake after that white frame."),
 ("stunTurns",1,"Turns a unit is stunned for once its last layer breaks. It skips its line and regrows nothing."),
 ("lineTravelMs",200,"How long the line takes to slide the next station to the centre."),
 ("lineFlashMs",90,"How long a centred station flashes before it strikes."),
 ("chargeShownMax",2,"Most charge stations drawn in an ability box; beyond this it just repeats."),
 ("critChance",0.05,"Default chance an attack crits, when the ability does not say otherwise."),
 ("frameMaxW",420,"The phone frame's widest; nothing drawn for it may exceed this."),
 ("momentPx",2,"Upscale of the title screen's mood line: bigger = chunkier pixels."),
 ("momentSize",10,"LOGICAL type size; what you see is this times momentPx."),
 ("critTagMs",700,"How long a plain positioned tag stays up."),
 ("critHoldMs",1000,"How long CRITICAL is held before the earned slot flies in."),
 ("critFlashMs",260,"One wash of the screen in the attacking emotion's colour."),
 ("critSlotFlyMs",520,"The earned slot travelling from the tag to the line."),
 ("critMult",2,"Damage multiplier on a critical."),
 ("statusStepMs",2000,"A station that applies a status holds this long, so it can be read."),
 ("attackStepMs",600,"Total time one executed station takes, start to finish."),
 ("lineAlertFlashes",3,"How many times a line flashes white before it starts executing."),
 ("lineAlertMs",110,"Length of one of those flashes."),
 ("unsheathMs",340,"How long one interface element takes to draw itself into place."),
 ("unsheathGapMs",110,"Gap before the next element starts — smaller than unsheathMs, so they overlap."),
 ("longPressMs",420,"How long an ability must be held before its tooltip opens."),
 ("swipeMinPx",36,"How far a pointer must travel across the panel to count as a page swipe."),
 ("tagGrowth",0.85,"How much an accumulating tag grows once it holds a full max-MS worth of change."),
 ("tagBreathSlowMs",1700,"Breathing period of an accumulating tag holding almost nothing."),
 ("tagBreathFastMs",380,"Breathing period of a tag holding a full max-MS worth — it flashes far harder."),
 ("gaugeW",134,"Gauge canvas width in logical pixels. Lower = chunkier pixels, coarser value steps."),
 ("gaugeH",22,"Gauge canvas height in logical pixels — tall enough for the end caps."),
 ("barCoreH",12,"Height of the bar's interior, before the wave, in logical pixels."),
 ("barWaveAmp",1.4,"Wave amplitude on the bar's silhouette (reference: 9.8 of a 1026-long bar)."),
 ("barWaveHalf",5.6,"Half-period of the silhouette wave, i.e. one arc (reference: 43.3)."),
 ("barWaveSpeed",0.10,"How fast the silhouette wave travels."),
 ("ecBandFrac",0.82,"Charge fill height as a fraction of the bar interior (reference: 76 of 92.74)."),
 ("ecWaveAmp",0.7,"Amplitude of the charge fill's own wave — long and shallow, per the reference."),
 ("ecWaveHalf",19,"Half-period of the charge fill's wave. Much longer than the silhouette's."),
 ("ecWaveSpeed",0.05,"How fast the charge wave travels."),
 ("ecInnerFrac",0.5,"Inner charge bar thickness, as a fraction of the outer one. Same shape, dodged over it."),
 ("ecInnerAmt",0.4,"How hard the inner charge bar dodges. 1 blows out to white; lower keeps the hue."),
 ("arcSpacing",3.1,"Spacing of the background arc texture (reference: 23.83)."),
 ("arcRadius",12.6,"Radius of the background arcs (reference: 96.95)."),
 ("arcAlpha",0.11,"How strongly the arcs show. They only read against the unlit stretch."),
 ("capW",5,"End-cap width in logical pixels."),
 ("capR",2,"End-cap corner rounding in logical pixels."),
 ("ballFlyMs",330,"How long a ball of light takes to arc from its indicator into the bar."),
 ("settleStepMs",70,"Pause between one hit landing on a bar and the next setting off."),
 ("fnumMinPx",20,"Smallest a floating damage number gets."),
 ("fnumMaxPx",62,"Largest — reached when one hit costs a whole stamina bar."),
 ("fnumRisePx",34,"How far a floating number travels up before it settles."),
 ("fnumInMs",600,"Fade in while rising."),
 ("fnumHoldMs",600,"Easing to a stop."),
 ("fnumOutMs",400,"Fade out."),
 ("fnumLingerMs",300,"Extra time a floating number holds before it goes."),
 ("hitShards",3,"Shapes thrown off where an attack lands."),
 ("hitShardMs",260,"How long one of those shapes flashes."),
 ("actionLineMs",620,"Buff / debuff streaks over the target."),
 ("statusFxMs",1000,"Distortion and bloom when a status takes hold."),
 ("backdropGlow",0.55,"How strongly the enemy's emotion tints the battle backdrop."),
 ("barTweenMs",1000,"How long a bar takes to travel to its new value once a hit lands."),
 ("overloadHoldMs",2000,"How long the OVERLOAD tag and its slow motion last."),
 ("ecScrollSpeed",0.0044,"How fast the charge gradient drifts along the bar."),
 ("themeOpening","audio/theme-opening.wav","Theme part 1: plays once."),
 ("themeLoop","audio/theme-loop.wav","Theme part 2: loops for ever, scheduled to start the instant part 1 ends."),
 ("musicFadeMs",900,"How quickly the theme fades out when the battle ends."),
 ("musicVolume",1.60,"Theme level, and ABOVE 1 on purpose: this is a gain, not a percentage. The battle theme is mastered about 10 dB quieter than the map's two, so it needs lifting to sit level with them - measured busy RMS -20.1 dBFS against -9.2 and -10.6, all three now landing at -16.0. Its peak is -5.2 dBFS, so 1.6x leaves headroom and cannot clip. Lower this if the sound effects are getting buried."),
 ("sfxVolume",1.05,"Sound-effect level, mixed against musicVolume."),
 ("layerEase",0.16,"How fast layers slide to their new slot. Higher = snappier."),
 ("layerWaveDelay",0.85,"Phase offset per slot, so breathing travels outward in waves."),
 ("layerInnerShrink",0.85,"Curve of the radius falloff toward the centre."),
 ("layerFill",0.42,"Ring thickness as a fraction of the gap to the next slot."),
 ("layerOuterThick",2,"Thickness multiplier for the outermost (active) layer."),
 ("layerInnerThick",0.9,"Thickness multiplier for every layer behind it."),
 ("layerFlashMs",50,"How long a struck layer flashes white before vanishing."),
 ("layerGapMs",18,"Beat between the layer vanishing and it regrowing at the back."),
 ("layerRegrowMs",380,"How long the regrow beat holds before resolution continues."),
 ("enemyRingBaseR",28,"Enemy: radius of the outermost ring, in logical canvas pixels."),
 # ---- what a TIER looks like. Nothing here is a stat; it is all silhouette. ----
 ("enemyShapeWeak","TRIANGLE","Battle silhouette for a WEAK enemy: its layer rings are concentric TRIANGLES."),
 ("enemyShapeRegular","CIRCLE","Battle silhouette for a REGULAR enemy. The original, and still the default for any tier the sheet does not name."),
 ("enemyShapeStrong","STAR","Battle silhouette for a STRONG enemy: concentric stars of enemyStarPoints points."),
 ("enemyStarPoints",7,"Points on a STRONG enemy's star."),
 ("enemyStarInner",0.62,"A star's valley radius as a fraction of its point radius. LOWER = narrower, sharper points; HIGHER = a blunter, rounder star. Below about 0.45 the thin inner rings start breaking up at this canvas size."),
 ("enemyShapeFill",1.6,"Band thickness multiplier for a SHAPED enemy only (triangle or star), on top of layerFill. A shape squeezes every band's pixel thickness by how narrow it is at that angle, and at the default the thin inner rings fell under one pixel and broke into dashes. THERE IS A CEILING AS WELL AS A FLOOR: at 2.0 the gaps between bands drop under a pixel and the rings merge into one solid shape, which is a filled triangle rather than concentric ones. Circles are not affected."),
 ("enemyShapeBreathe",0.6,"How much of the usual ring breathing a SHAPED enemy does. Full amplitude closes the gap between two neighbouring bands, and with the thicker bands above that swallows one of them. Together with enemyShapeFill this is what keeps every ring continuous at every angle - see tools/check_rings.js."),
 ("enemyTriRound",0.72,"How far a WEAK enemy's triangle goes toward a TRUE triangle: 1 is sharp corners, 0 is a circle. It is not there to soften the look - a true triangle squeezes the thin inner rings under one pixel near the flat sides and they disappear."),
 ("enemyShapeSpin",0.018,"Radians per frame a triangle or a star turns. A polygon that never moves reads as a logo; a slow turn reads as a creature. 0 stops it."),
 ("enemyRingScaleWeak",0.85,"Overall size of a WEAK enemy in battle, against a REGULAR one."),
 ("enemyRingScaleStrong",1.05,"Overall size of a STRONG enemy in battle. Do not raise much: the canvas is 64px and the rings breathe outward past this."),
 ("rideScaleWeak",0.78,"How big a WEAK enemy is during a ride, against its element row's own size. Do not go much lower: below about nine pixels across, three points stop reading as a triangle and it is just a small blob."),
 ("rideScaleStrong",1.55,"How big a STRONG enemy is during a ride. Size is the tier cue that survives being glanced at; the silhouette is the one that survives being looked at."),
 ("enemyRingSpacing",4.7,"Enemy: radius step between one layer and the next in. baseR/spacing sets how many fit."),
 ("enemyRingBreathe",1.6,"Enemy ring breathing amplitude."),
 ("playerRingBaseR",300,"Player: radius of the outermost ring. Huge on purpose — a big circle reads as a nearly flat band."),
 ("playerRingSpacing",5,"Player: radius step between layers. Small, so all six sit in a thin arc despite the huge radius."),
 ("playerRingBreathe",2.2,"Player ring breathing amplitude."),
 ("playerCanvasW",120,"Player ring canvas width in logical pixels (wide and short)."),
 ("playerCanvasH",32,"Player ring canvas height in logical pixels."),
 ("playerRingCy",300,"Player: circle centre Y, pushed far below the strip so only the flat top of the arc shows."),
 ("crusherSteps",12,"Audio bit-crusher quantisation. Low = Atari harsh, high = SNES-clean."),
 # ---- MAP: the ride ----
 ("travelRollSecs",3,"Legacy fixed roll interval. rollMinSecs/rollMaxSecs supersede it."),
 ("travelTarget",10,"Legacy fixed segment target. segBase x diltransience supersedes it."),
 ("travelTapR",15,"FLOOR for how near a tap must land on a travel element, in art pixels. The real reach is the element's own size scaled up; this is the minimum."),
 ("segBase",5,"Track Segments to bridge a link, before diltransience and world state."),
 ("rollMinSecs",1,"Fastest the track event roll can come round (navigation GDD 4: every 1-3s)."),
 ("rollMaxSecs",3,"Slowest the track event roll can come round. Enemy density moves it."),
 ("aggroLockSecs",5,"Default countdown before an aggro enemy forces the encounter."),
 ("travelMaxLive",5,"How many elements may be out at once, ALL KINDS TOGETHER. No roll happens while the board is full, so it has to clear before anything new appears."),
 ("tripFillMs",520,"How long the Emotional Trip bar takes to ease across to a newly collected segment."),
 ("tripFlashMs",260,"How long the bar flashes white when a segment lands in it."),
 ("departSecs",3,"How long the camera spends zooming into the departure station before the wipe opens."),
 ("leanEaseMs",520,"How long the map takes to tilt into, and back out of, the focused-station lean."),
 ("leanDeg",20,"How far the map tilts when a destination is being considered."),
 ("camEaseMs",620,"How long the camera takes to travel to a station the player has just picked."),
 # ---- MAP <-> BATTLE: the bridge ----
 ("orbMsPct",0.18,"How much of MaxMS one Stamina Orb gives back. Orbs only matter mid-ride, which is the only time MS moves."),
 ("encounterWashSecs",1.1,"How long the travel screen floods with the enemy's colour before the wipe onto the fight."),
 ("battleWipeMs",820,"The circular wipe that opens onto the battle."),
 ("battleFadeMs",620,"The fade to white that unloads the battle and gives the ride back."),
 ("musicCutMs",120,"The ultra-quick fade that kills the battle theme the moment an enemy falls."),
 ("enterWipeMs",1100,"The circular wipe that opens the map when arriving from the title screen."),
 # ---- MAP: the wager ----
 ("exitKeepPct",0.6,"Crystals kept when EXITING at an intermediate platform."),
 ("defeatKeepPct",0.1,"Crystals kept on defeat."),
 ("itemLossOnExit",0.4,"Chance each unbanked item is lost when exiting early."),
 ("itemLossOnDefeat",0.9,"Chance each unbanked item is lost on defeat."),
 # ---- MAP: the world ----
 ("fogPeekHide",1.6,"Fog at or above this hides a station's details from the map."),
 ("startLineKeys","L1|L2","Line Keys the player begins with."),
 # ---- the player's profile and kit ----
 ("affinitySlots",2,"How many Emotional Affinities a player picks at creation."),
 ("armorSlots",1,"Emotional Armor pieces worn at once."),
 ("startArmor","ARM_SCARS","The armor a new profile begins in."),
 ("startSets","LO_ANGER|LO_SADNESS|LO_JOY","The Move Sets a new profile begins with."),
]
sheet("rules", ["key","value","description"], {"key","value"},
  [dict(key=k, value=v, description=d) for k,v,d in RULES],
  widths={"key":24,"value":14,"description":78},
  notes="Global tunables. Keys are read by name — renaming one breaks whatever reads it.")

# ───────────────────────────── sounds ─────────────────────────────
SND = [
 ("tap","square",980,1460,80,0.46,0.30,"Adding an ability to the line."),
 ("place","square",620,1760,130,0.40,0.35,"The station lands on your line and flashes white."),
 ("resist","square",880,760,120,0.34,0.30,"A layer shrugs off its own emotion. Played twice, descending."),
 ("hit_anger","square",300,90,150,0.36,0.28,"Anger lands: hard and low."),
 ("hit_surprise","sawtooth",1200,300,130,0.30,0.34,"Surprise lands: a sharp drop."),
 ("hit_disgust","noise",420,200,180,0.30,0.30,"Disgust lands: wet and dull."),
 ("hit_joy","square",900,1800,120,0.30,0.32,"Joy lands: bright and rising."),
 ("hit_sadness","triangle",520,180,220,0.30,0.38,"Sadness lands: a long fall."),
 ("hit_fear","noise",900,1400,120,0.26,0.40,"Fear lands: a cold hiss."),
 ("buff_up","square",420,1500,260,0.28,0.35,"Something is lifted."),
 ("debuff_down","sawtooth",1500,260,300,0.30,0.38,"Something is dragged down."),
 ("status_on","noise",200,1100,520,0.32,0.5,"A status takes hold."),
 ("sheatheE","sawtooth",1500,220,150,0.30,0.22,"Enemy UI element drawn into place. Lower."),
 ("sheatheP","sawtooth",2100,340,140,0.28,0.22,"Player UI element drawn into place. Higher."),
 ("slot","triangle",300,1500,150,0.30,0.45,"A temporary slot flies in from the opponent."),
 ("remove","square",520,220,90,0.34,0.24,"Removing a station."),
 ("depart","sawtooth",240,70,420,0.26,0.45,"The line departs and charge is spent."),
 ("travel","square",150,150,38,0.10,0.15,"Line scrolling to the next station."),
 ("arrive","square",880,880,45,0.14,0.30,"Station reaches the head and lights up."),
 ("hit","noise",700,700,190,0.34,0.40,"Off-type hit."),
 ("absorb","square",300,1200,260,0.24,0.50,"Same-type hit: the layer absorbs it."),
 ("block","noise",1600,1600,90,0.30,0.35,"Shield blocks, or shield goes up."),
 ("breaklayer","sawtooth",760,80,340,0.28,0.55,"A layer flashes and breaks."),
 ("regrow","triangle",120,520,300,0.16,0.40,"The broken layer regrows at the back."),
 ("clash","noise",240,240,1500,0.40,0.65,"Death: the chevrons meet."),
 ("speak","square",620,540,42,0.09,0.12,"One letter of dialogue appearing."),
 ("win","square",440,1320,620,0.24,0.55,"Victory."),
 ("lose","sawtooth",320,50,950,0.28,0.60,"Defeat."),
 # ---- MAP: travelling between stations ----
 ("map_select","square",760,1180,70,0.30,0.25,"MAP: an adjacent station is chosen."),
 ("map_depart","sawtooth",180,900,620,0.30,0.50,"MAP: the wipe opens and the train pulls away."),
 ("map_clack","square",210,150,34,0.13,0.18,"MAP: one rail joint passing under the train."),
 ("map_announce","triangle",520,1040,300,0.26,0.40,"MAP: 'Propera parada' over the carriage speaker."),
 ("map_flash","noise",300,2600,700,0.34,0.55,"MAP: the white flood, the train leaving the frame."),
 ("map_arrive","square",660,1980,540,0.26,0.45,"MAP: the station banner lands."),
 ("map_return","sawtooth",1300,260,560,0.24,0.42,"MAP: the wipe closes back onto the map."),
 ("map_step","triangle",900,1350,90,0.20,0.22,"MAP: the player marker moving one station."),
 ("map_pick","square",1180,1760,70,0.26,0.20,"MAP: a Track Segment is grabbed."),
 ("map_collect","square",880,2200,150,0.28,0.35,"MAP: it lands in the Emotional Trip bar."),
 ("map_entity","noise",520,240,180,0.28,0.45,"MAP: an Emotional Entity is prodded."),
 # ---- MAP: the interface itself ----
 ("ui_tap","square",900,1240,52,0.22,0.14,"MAP: a button."),
 ("ui_open","square",520,1180,120,0.24,0.28,"MAP: a panel comes up."),
 ("ui_close","square",1180,520,110,0.20,0.22,"MAP: a panel goes away."),
 ("ui_page","square",760,1010,60,0.18,0.16,"MAP: moving between tabs."),
 ("ui_equip","square",640,1560,170,0.26,0.32,"MAP: something is put on."),
 ("ui_deny","sawtooth",340,180,150,0.24,0.20,"MAP: refused - not owned, no route, no key."),
 ("map_tripup","sine",520,880,180,0.22,0.18,"A segment lands in the Emotional Trip bar. Soft and rising — the one sound in the ride that is a reward rather than a click."),
 ("ui_station","triangle",680,1180,140,0.24,0.34,"MAP: a station panel opens."),
]
sheet("sounds", ["id","wave","f0","f1","dur","gain","echo","description"],
  {"id","wave","f0","f1","dur","gain","echo"},
  [dict(id=i,wave=w,f0=a,f1=b,dur=d,gain=g,echo=e,description=n) for i,w,a,b,d,g,e,n in SND],
  widths={"description":42},
  notes=("Every sound is synthesised at runtime — there are no audio files. "
         "wave: square | sawtooth | triangle | sine | noise. f0/f1 sweep in Hz over 'dur' ms. "
         "'echo' is how much of the sound feeds the shared delay line. Everything runs through a "
         "bit-crusher, which is where the crunch comes from."))

# ───────────────────────────── dialogue ─────────────────────────────
import csv as _csv
_dlg_src = os.path.join(APP, "sources", "neuro_metro_avui_enemy_dialogues.csv")
_dlg = []
if os.path.exists(_dlg_src):
    for r in _csv.DictReader(open(_dlg_src, encoding="utf-8-sig")):
        _dlg.append(dict(emotion=r["Emotion"].strip().upper(),
                         persona=r["Persona"].strip(),
                         state=r["State"].strip().upper(),
                         line=r["Dialogue Line"].strip(),
                         tier="", enabled=1, notes=""))

# ---- TIER-SCOPED PERSONAS ---------------------------------------------------
# An enemy picks a persona at random from the rows matching its EMOTION, which is
# what makes the same fight different twice. That alone would have The Enforcer
# speaking the Commuter's lines, though, since both are Anger — a strong enemy
# has to sound like one. So a row may name a `tier`, and the rule is:
#
#     rows for this emotion AND this tier   ->   use those
#     none                                  ->   fall back to the tier-blank rows
#
# The thirty imported rows stay blank and so stay available to everyone, and a
# WEAK or STRONG enemy speaks only in its own voice. Blank is not "no tier", it
# is "any tier".
def dlg(emotion, persona, tier, intro, winning, losing, defeat):
    states = [("INTRO", intro), ("WINNING", winning), ("LOSING", losing), ("DEFEAT", defeat)]
    return [dict(emotion=emotion, persona=persona, state=s, line=l, tier=tier,
                 enabled=1, notes="") for s, l in states]

for rows in [
 # ---- STRONG / ANGER: anger that has the upper hand, and knows it -----------
 dlg("ANGER", "The Bailiff", "STRONG",
     "I have the paper, I have the locks and I have all morning. OPEN IT.",
     "Everything in here belongs to someone else now. Including you.",
     "I'm just doing my job — I'M JUST DOING MY JOB!",
     "There's another twelve on the list today. Someone else will come."),
 dlg("ANGER", "The Riot Line", "STRONG",
     "Disperse. You have been warned. YOU HAVE BEEN WARNED!",
     "Nobody filmed this. Nobody ever films this.",
     "Hold the line... hold the LINE...",
     "Badge number... nobody ever asks for the badge number."),
 dlg("ANGER", "The Developer", "STRONG",
     "Forty families out by Friday and a rooftop pool by spring. Try and stop me.",
     "This whole barri is already sold. You're standing in a render.",
     "The permits — who talked to the press about the permits?!",
     "It was going to be beautiful..."),
 # ---- STRONG / SURPRISE: the thing that turns the day over -----------------
 dlg("SURPRISE", "The Verdict", "STRONG",
     "All of it. They're keeping ALL of it. I signed — I SIGNED WHAT?!",
     "Read the small print, they said. READ IT NOW, GO ON!",
     "There has to be an appeal. There's always an appeal...",
     "Fourteen years. In one afternoon."),
 dlg("SURPRISE", "The Blackout", "STRONG",
     "Whose jacket is this? WHOSE BLOOD IS THIS? Somebody tell me what I did!",
     "You were there! You saw it! SAY WHAT YOU SAW!",
     "It's coming back... oh God, it's coming back...",
     "Don't tell me. Please. Don't ever tell me."),
 dlg("SURPRISE", "The Collapse", "STRONG",
     "The ceiling was there this morning. IT WAS THERE THIS MORNING!",
     "Nothing holds! Look at it! NOTHING HOLDS!",
     "My hands won't stop. Why won't my hands stop?",
     "Everyone kept saying it was fine."),
 # ---- WEAK / SADNESS: too tired to be dangerous ----------------------------
 dlg("SADNESS", "The Sleepless", "WEAK",
     "Fourth night. The train's warm, that's all. Don't make me move.",
     "You get used to the noise. You never get used to the light.",
     "I only wanted to close my eyes...",
     "Wake me at Fondo. Somebody always does."),
 dlg("SADNESS", "The Unread Message", "WEAK",
     "Delivered. Two blue ticks, eleven days ago. That's the whole story.",
     "Everyone's busy. Everyone's always so busy.",
     "Maybe she's just... maybe the phone...",
     "I'll type it again tonight. I won't send it."),
 dlg("SADNESS", "The Last One Out", "WEAK",
     "Everybody left for Berlin, Lisbon, anywhere. I locked up.",
     "You'll go too. They all go.",
     "It's only me on this platform now...",
     "Somebody has to stay and turn the lights off."),
 # ---- WEAK / JOY: joy with nothing behind it -------------------------------
 dlg("JOY", "The Last Round", "WEAK",
     "Six in the morning and the night's still GOING! Come on, one more, ONE MORE!",
     "This is living! THIS IS LIVING! Where's my phone?",
     "Wait — wait, has anyone got water...",
     "I'll be fine. I've got work at nine. I'll be fine."),
 dlg("JOY", "The Hen Party", "WEAK",
     "Eleven of us, matching sashes, and NOBODY is going home sad tonight!",
     "Smile! SMILE! It's meant to be the best day of her life!",
     "Where's Marta? Has anyone seen Marta?",
     "She cried in the toilets for an hour. We got the photos though."),
 dlg("JOY", "The Busker", "WEAK",
     "Same four chords for nine years and the carriage still claps! Listen!",
     "They're LOVING this! Look at them! They're loving it!",
     "Two euros. Whole carriage. Two euros...",
     "Next stop's better. Next stop's always better."),
]:
    _dlg.extend(rows)
# ───────────────────────────── loadouts ─────────────────────────────
# One Loadout per emotion. Slots are POSITIONAL — an empty cell is a real, visible
# empty slot in the panel, not a missing entry — so they are four columns rather
# than one pipe list. A future equip screen maps one-to-one onto them.
# A Loadout is what the progression GDD calls a MOVE SET. Same row, two names —
# the sheet keeps the old one so nothing in the battle system has to move.
LO_COLS = ["id","emotion","name","slot1","slot2","slot3","slot4",
           "tier","ec_mod","passive","cost","trade_in","enabled","notes"]
LO_LIVE = {"id","emotion","name","slot1","slot2","slot3","slot4",
           "tier","ec_mod","passive","cost","trade_in","enabled"}
def lo(i, emo, name, *slots, notes="", tier=1, ec_mod=0, passive="", cost="", trade_in=""):
    row = dict(id=i, emotion=emo, name=name, enabled=1, notes=notes,
               tier=tier, ec_mod=ec_mod, passive=passive, cost=cost, trade_in=trade_in)
    for n in range(4):
        row["slot%d" % (n+1)] = slots[n] if n < len(slots) else ""
    return row
LOADOUTS_ROWS = [
 lo("LO_ANGER",   "ANGER",   "Anger",   "ATK_ANGER","HVY_ANGER","GEN_ANGER"),
 lo("LO_SADNESS", "SADNESS", "Sadness", "ATK_SADNESS","HVY_SADNESS","INFLICT_SAD"),
 lo("LO_JOY",     "JOY",     "Joy",     "ATK_JOY","HVY_JOY","GEN_JOY"),
 lo("LO_DISGUST", "DISGUST", "Disgust", "GEN_DISGUST","ROT"),
 lo("LO_SURPRISE","SURPRISE","Surprise","ATK_SURPRISE","HVY_SURPRISE","STARTLE"),
 lo("LO_FEAR",    "FEAR",    "Fear",     notes="No Fear abilities exist yet."),
]
sheet("loadouts", LO_COLS, LO_LIVE, LOADOUTS_ROWS,
  widths={"notes":40},
  notes=("One Loadout per emotion; the player carries `equippedSlots` of them into a battle "
         "(units.loadouts). A Loadout may hold any ability whose emotion set includes its own "
         "emotion — see abilities.emotions for hybrids. Blank slots render as empty."))

# ───────────────────────────── prompts ─────────────────────────────
# The line's cue. One is drawn at random each turn, so the game keeps asking the
# question in a slightly different voice.
PROMPTS = [
 "Whatcha gonna do?", "How did that make you feel?", "Act!", "R U OK?",
 "Show them.", "Say something.", "Your move.", "Don't just stand there.",
 "Feel it.", "What now?", "Answer that.", "Let it out.",
 "Breathe. Then act.", "You alright?", "Do something.",
]
sheet("prompts", ["id","text","enabled","notes"], {"id","text","enabled"},
  [dict(id="p%02d" % (i+1), text=t, enabled=1, notes="") for i,t in enumerate(PROMPTS)],
  widths={"text":40},
  notes="The cue above the attack line. One is picked at random each turn.")

# ───────────────────────────── moments ─────────────────────────────
# The title screen names Barcelona's current moment. `day` is a weekday or "*",
# and from_hour/to_hour are BARCELONA local hours, to_hour exclusive; a band that
# wraps past midnight (22 -> 5) is written as from=22 to=5. Higher priority wins,
# so a Friday-night line beats the everyday one; ties are picked at random.
MOM_COLS = ["id","day","from_hour","to_hour","phrase","priority","enabled","notes"]
MOM_LIVE = {"id","day","from_hour","to_hour","phrase","priority","enabled"}
def mo(i, day, a, b, phrase, pri=0):
    return dict(id=i, day=day, from_hour=a, to_hour=b, phrase=phrase,
                priority=pri, enabled=1, notes="")
MOMENTS = [
 # ---- everyday, by band ----
 mo("dawn_1","*",5,7,"the first light is still deciding"),
 mo("dawn_2","*",5,7,"nobody has spoken yet today"),
 mo("morn_1","*",7,10,"early morning"),
 mo("morn_2","*",7,10,"the early carriages are full of unfinished sleep"),
 mo("morn_3","*",7,10,"everyone is already late"),
 mo("mid_1","*",10,13,"middle of the day"),
 mo("mid_2","*",10,13,"the light is flat and honest"),
 mo("noon_1","*",13,16,"all afternoon ahead"),
 mo("noon_2","*",13,16,"the long slow part of the day"),
 mo("late_1","*",16,19,"the afternoon is running out"),
 mo("late_2","*",16,19,"everyone is going somewhere they did not choose"),
 mo("dusk_1","*",19,21,"somber twilight"),
 mo("dusk_2","*",19,21,"the hour the day admits what it was"),
 mo("night_1","*",21,24,"the night is getting loud"),
 mo("night_2","*",21,24,"the last trains are filling up"),
 mo("small_1","*",0,5,"the madrugada, and still moving"),
 mo("small_2","*",0,5,"too late to be anything but honest"),
 # ---- weekday specifics ----
 mo("mon_morn","MON",5,10,"monday, and the week already weighs something",2),
 mo("mon_dusk","MON",19,24,"the longest monday of the month",2),
 mo("wed_mid","WED",10,16,"the week is stuck exactly in the middle",2),
 mo("thu_late","THU",16,21,"almost, but not yet",2),
 mo("fri_dusk","FRI",19,21,"a night to finally be free",2),
 mo("fri_night","FRI",21,24,"a night to finally be free",2),
 mo("fri_small","FRI",0,5,"friday refusing to end",2),
 mo("sat_mid","SAT",10,16,"a saturday with nothing owed to anyone",2),
 mo("sat_night","SAT",21,24,"the city is pretending it never has to work again",2),
 mo("sun_mid","SUN",10,16,"a sunday that stretches too far",2),
 mo("sun_dusk","SUN",19,24,"sunday dusk, and monday already breathing on it",2),
]
sheet("moments", MOM_COLS, MOM_LIVE, MOMENTS,
  widths={"phrase":54,"notes":30},
  notes=("Title-screen mood line: \"Barcelona, <weekday>, <phrase>.\" day is MON..SUN "
         "or * for any. Hours are Barcelona local, to_hour exclusive; a band may wrap "
         "past midnight. Highest priority wins, ties are random."))

sheet("dialogue", ["emotion","persona","state","line","tier","enabled","notes"],
  {"emotion","persona","state","line","tier"}, _dlg,
  widths={"line":92,"persona":20,"tier":10,"notes":20},
  notes=("What each enemy says. states: INTRO (battle start) - WINNING (player drops below 20% MS) - "
         "LOSING (this enemy drops below 20% MS) - DEFEAT (this enemy dies). A battle picks one persona "
         "at random from the rows matching the enemy's emotion, so the same enemy type speaks differently "
         "each run. `tier` narrows that: a WEAK or STRONG enemy uses only the rows carrying its own tier, "
         "and everything else uses the rows where tier is BLANK. Blank means ANY tier, not 'no tier'. "
         "Give a new persona all four states or it will fall silent at the moment it is missing."))

# ───────────────────────────── checks ─────────────────────────────
ck = WB.create_sheet("checks")
ck["A1"] = "LIVE VALIDATION — every row should read OK. Anything else means a broken reference."
ck["A1"].font = NOTE_FONT
hdr = ["check","expected","actual","status"]
for c,h in enumerate(hdr,1):
    x = ck.cell(row=3, column=c, value=h); x.font = HDR_FONT; x.fill = HDR_FILL; x.border = BORD
CHECKS = [
 ("Abilities with an emotion not in 'emotions'", 0,
  '=SUMPRODUCT((abilities!C5:C200<>"")*(COUNTIF(emotions!A5:A200,abilities!C5:C200)=0))'),
 ("Abilities missing an id", 0, '=SUMPRODUCT((abilities!B5:B200<>"")*(abilities!A5:A200=""))'),
 ("Units with a blank pool", 0, '=SUMPRODUCT((units!A5:A200<>"")*(units!H5:H200=""))'),
 ("Matchup rows with no label", 0, '=SUMPRODUCT((matchups!A5:A200<>"")*(matchups!F5:F200=""))'),
 ("Matchup attack emotions unknown (blank/*/NONE allowed)", 0,
  '=SUMPRODUCT((matchups!A5:A200<>"")*(matchups!A5:A200<>"*")*(matchups!A5:A200<>"NONE")*(COUNTIF(emotions!A5:A200,matchups!A5:A200)=0))'),
 ("Matchup layer emotions unknown (blank/*/NONE allowed)", 0,
  '=SUMPRODUCT((matchups!B5:B200<>"")*(matchups!B5:B200<>"*")*(matchups!B5:B200<>"NONE")*(COUNTIF(emotions!A5:A200,matchups!B5:B200)=0))'),
 ("Duplicate ability ids", 0,
  '=SUMPRODUCT((abilities!A5:A200<>"")*(COUNTIF(abilities!A5:A200,abilities!A5:A200&"")>1))'),
 ("Duplicate emotion ids", 0,
  '=SUMPRODUCT((emotions!A5:A200<>"")*(COUNTIF(emotions!A5:A200,emotions!A5:A200&"")>1))'),
 ("Unit emotions not in 'emotions' (player is blank)", 0,
  '=SUMPRODUCT((units!C5:C200<>"")*(COUNTIF(emotions!A5:A200,units!C5:C200)=0))'),
 ("Unit tiers that are not WEAK/REGULAR/STRONG (blank allowed)", 0,
  '=SUMPRODUCT((units!D5:D200<>"")*(units!D5:D200<>"WEAK")*(units!D5:D200<>"REGULAR")*(units!D5:D200<>"STRONG"))'),
 ("Statuses an ability applies that do not exist", 0,
  '=SUMPRODUCT((abilities!AA5:AA200<>"")*(COUNTIF(status_effects!A5:A200,abilities!AA5:AA200)=0))'),
]
for r,(label,exp,formula) in enumerate(CHECKS, 4):
    ck.cell(row=r, column=1, value=label).font = BODY
    ck.cell(row=r, column=2, value=exp).font = BODY
    ck.cell(row=r, column=3, value=formula).font = BODY
    ck.cell(row=r, column=4, value=f'=IF(C{r}=B{r},"OK","CHECK")').font = BOLD
    for c in range(1,5): ck.cell(row=r, column=c).border = BORD
ck.column_dimensions["A"].width = 56
for col,w in (("B",12),("C",12),("D",12)): ck.column_dimensions[col].width = w

# ─────────────────── MAP: stations + metro_lines ───────────────────
# Seeded from the CSVs rather than inlined: the network is 110 stations and
# would bury the rest of this generator. Re-exporting from Sheets overwrites
# them, which is the intended direction of travel.
def _seed(table):
    """Read a sheet's rows back out of its exported CSV.

    THIS READS ITS OWN OUTPUT, so it must survive the banner rows `sheet()`
    writes. A plain DictReader does not: on the second round-trip it took the
    notes banner as the header, which collapsed every seeded sheet to one
    column and silently destroyed stations, lines, elements, items, world bands
    and armor. Find the real header the way build_data.clean() does — the first
    row whose first cell is a key column — and drop the LIVE/planned tag row
    under it."""
    import csv as _csv
    path = os.path.join(csv_dir(), f"battle-system-config - {table}.csv")
    if not os.path.exists(path):
        print(f"  (no seed CSV for {table}, sheet left empty)"); return [], []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        raw = [r for r in _csv.reader(fh) if any(c.strip() for c in r)]
    if not raw:
        return [], []
    keys = {"id", "key", "attack_emotion", "emotion"}
    start = next((i for i, r in enumerate(raw) if r and r[0].strip() in keys), 0)
    head = [h.strip() for h in raw[start]]
    body = raw[start + 1:]
    if body and body[0] and body[0][0].strip().upper() in ("LIVE", "PLANNED"):
        body = body[1:]
    if len(head) < 2:
        raise SystemExit(f"! {table}: header looks wrong ({head!r}) — refusing to "
                         f"rebuild the workbook from a damaged CSV")
    out = []
    for r in body:
        r = (r + [""] * len(head))[:len(head)]
        out.append(dict(zip(head, r)))
    return head, out

ST_COLS, ST_ROWS = _seed("stations")
if ST_COLS:
    sheet("stations", ST_COLS, {"id","name","lines","emotions","state","spawn","fog","threat","diltransience",
           "x","y","enabled"}, ST_ROWS,
          widths={"id":30,"name":30,"lines":14,"emotions":22,"state":14,"spawn":22,"notes":30},
          notes="One row per PLACE, not per stop — an interchange is a single station several "
                "lines call at. `state` is the station's condition (blank = nothing special); it "
                "drives what the map shows and what happens to a player passing through, in either "
                "direction. x/y are schematic world units; the renderer squares every link to 45 "
                "degrees, so they only need to be roughly right.")

AR_COLS, AR_ROWS = _seed("armor")
if AR_COLS:
    sheet("armor", AR_COLS,
          {"id","name","tier","ms_mod","layer1","layer2","passive","cost","trade_in","enabled"},
          AR_ROWS, widths={"id":16,"name":24,"passive":16,"cost":24,"trade_in":14,"notes":46},
          notes="Emotional Armor. `ms_mod` is ADDED to the player's base MaxMS; layer1/layer2 are "
                "the Emotional Layers it grants in battle (blank = none). One passive per piece. "
                "cost/trade_in are for the Wandering Store, which is not built yet.")

IT_COLS, IT_ROWS = _seed("items")
if IT_COLS:
    sheet("items", IT_COLS, {"id","name","emotion","weight","enabled"}, IT_ROWS,
          widths={"id":20,"name":20,"notes":46},
          notes="PLACEHOLDERS. Item design is not written yet — these exist so the roll-and-lose "
                "machinery has something to roll. `weight` is relative drop likelihood.")

WB_COLS, WB_ROWS = _seed("world_bands")
if WB_COLS:
    sheet("world_bands", WB_COLS,
          {"id","day","from_hour","to_hour","weather","fog","density","aggro","diltransience","enabled"},
          WB_ROWS, widths={"id":14,"day":26,"weather":14,"notes":42},
          notes="day / hour / weather -> multipliers on a station's live attributes. EVERY matching "
                "row multiplies, so BASE is the floor and the rest stack onto it. Hours are Barcelona "
                "local, to_hour exclusive, and a band may wrap past midnight. Weather comes from "
                "WorldState (stubbed today, a real API later).")

TE_COLS, TE_ROWS = _seed("travel_elements")
if TE_COLS:
    sheet("travel_elements", TE_COLS,
          {"id","name","kind","motion","chance","max_on_screen","max_per_trip","life_min",
           "life_max","size","drift","worth","payload","amount","lock_secs","unit","enabled"}, TE_ROWS,
          widths={"id":20,"name":18,"notes":46},
          notes="What can appear around the train during a ride. `chance` is rolled per element "
                "every travelRollSecs; max_on_screen caps how many exist at once, max_per_trip "
                "how many one ride may ever produce. life_min/life_max in seconds. `drift` is "
                "speed relative to the track (>1 reads as nearer the camera). These are BASES: "
                "the target station's `spawn` column scales them, and so will weather and time "
                "of day when those exist. `unit` FORCES which units row an enemy element fights "
                "as; LEAVE IT BLANK and the map rolls one from the units that name the line "
                "being ridden in their `spawn_lines`, which is what makes Line 2 feel unlike "
                "Line 5. Fill it in only to pin one element to one enemy.")

CS_COLS, CS_ROWS = _seed("city_status")
if not CS_COLS:
    CS_COLS = ["id","name","emotions","fx","lines","share","day","hours",
               "density","aggro","fog","diltransience","blurb","enabled","notes"]
    CS_ROWS = [dict(id="RUSH_HOUR", name="Rush Hour", emotions="DISGUST|ANGER", fx="RUSH",
                    lines="L1|L2|L5", share=0.4, day="MON|TUE|WED|THU|FRI",
                    hours="7-10|17-20", density=1.6, aggro=1.5, fog="", diltransience="",
                    blurb="The platforms are packed. Far more of them are out on the "
                          "affected stretches, and they are quicker to take an interest.",
                    enabled=1,
                    notes="The reference status, and the only one designed. Everything else "
                          "waits on the design.")]
sheet("city_status", CS_COLS,
      {"id","name","emotions","fx","lines","share","hours","enabled"}, CS_ROWS,
      widths={"id":16,"name":18,"emotions":20,"lines":16,"day":26,"hours":16,
              "blurb":56,"notes":46},
      notes="CITY-WIDE CONDITIONS. A status is something happening to Barcelona that picks a "
            "SUBSET of the stations on the lines it names and changes them: their attributes "
            "multiply by the columns here, and the map paints `fx` around them. `hours` holds "
            "one or more Barcelona-local windows (7-10|17-20), to_hour exclusive, and a window "
            "may wrap past midnight. `share` is the fraction of eligible stations affected - "
            "WHICH ones comes from a hash of the station, the status and the current window, so "
            "every client picks the same set and it holds still until the window ends. "
            "`emotions` drives the colours the effect is drawn in (a mix reads as a mix); "
            "`blurb` is what the tag says when it is tapped.")

ML_COLS, ML_ROWS = _seed("metro_lines")
if ML_COLS:
    sheet("metro_lines", ML_COLS, {"id","name","emotion","stations","enabled"}, ML_ROWS,
          widths={"id":8,"name":14,"emotion":14,"stations":90,"notes":26},
          notes="One row per line, `stations` in running order (| separated). Colour is NOT stored: "
                "it comes from the emotion, so the map cannot drift out of step with battle.")


# ───────────────────────────── the checks a FORMULA cannot make ─────────────────────────────
# Every cross-reference in this workbook that is a PIPE LIST — a unit's pool, its
# loadouts, its spawn_lines, a Loadout's four slots — is invisible to the `checks`
# sheet, because a spreadsheet cannot look inside "ATK_ANGER|HVY_ANGER|BLIND". They
# are exactly the references most likely to be wrong, since they are the ones typed
# by hand. So they are checked HERE, at build time, and a bad one stops the build.
#
# A dangling id does not crash anything at runtime: `pool.map(a=>ABILITIES[a]).filter(Boolean)`
# quietly drops it, and the enemy simply fights with one fewer ability than it was
# designed with. That is the worst kind of bug — a balance change nobody made.
def _validate():
    ab_ids  = {r["id"] for r in ABILS}
    emo_ids = {e[0] for e in EMO}
    st_ids  = {r["id"] for r in SE_ROWS}
    lo_ids  = {r["id"] for r in LOADOUTS_ROWS}
    line_ids = {r["id"] for r in ML_ROWS} if ML_COLS else set()
    bad = []
    def pipes(v): return [s.strip() for s in str(v or "").split("|") if s.strip()]
    for u in UNITS_ROWS:
        for a in pipes(u.get("pool")):
            if a not in ab_ids: bad.append(f"units.{u['id']}.pool -> no ability '{a}'")
        for l in pipes(u.get("loadouts")):
            if l not in lo_ids: bad.append(f"units.{u['id']}.loadouts -> no loadout '{l}'")
        for e in pipes(u.get("layers")):
            if e not in emo_ids: bad.append(f"units.{u['id']}.layers -> no emotion '{e}'")
        if u.get("emotion") and u["emotion"] not in emo_ids:
            bad.append(f"units.{u['id']}.emotion -> no emotion '{u['emotion']}'")
        if u.get("tier") not in ("", "WEAK", "REGULAR", "STRONG"):
            bad.append(f"units.{u['id']}.tier -> '{u['tier']}' is not WEAK/REGULAR/STRONG")
        for s in pipes(u.get("spawn_lines")):
            name = s.split(":")[0].strip()
            if name != "*" and line_ids and name not in line_ids:
                bad.append(f"units.{u['id']}.spawn_lines -> no line '{name}'")
            try: float(s.split(":")[1])
            except (IndexError, ValueError):
                bad.append(f"units.{u['id']}.spawn_lines -> '{s}' has no weight after the colon")
    for a in ABILS:
        if a.get("status_apply") and a["status_apply"] not in st_ids:
            bad.append(f"abilities.{a['id']}.status_apply -> no status '{a['status_apply']}'")
    for l in LOADOUTS_ROWS:
        for n in range(1, 5):
            v = l.get("slot%d" % n)
            if v and v not in ab_ids:
                bad.append(f"loadouts.{l['id']}.slot{n} -> no ability '{v}'")
    # Every persona has to be able to say all four things, or an enemy goes silent
    # at the one moment the fight was about to be interesting.
    seen = {}
    for r in _dlg:
        seen.setdefault((r["emotion"], r["persona"], r["tier"]), set()).add(r["state"])
    for (emo, per, tier), states in sorted(seen.items()):
        missing = {"INTRO", "WINNING", "LOSING", "DEFEAT"} - states
        if missing:
            bad.append(f"dialogue '{per}' ({emo}{'/'+tier if tier else ''}) has no "
                       + ", ".join(sorted(missing)))
    # An enemy whose emotion+tier finds no persona at all would fight nameless.
    for u in UNITS_ROWS:
        if u.get("tags") != "ENEMY": continue
        emo, tier = u.get("emotion"), u.get("tier") or ""
        pool = [k for k in seen if k[0] == emo and k[2] == tier]
        if not pool: pool = [k for k in seen if k[0] == emo and k[2] == ""]
        if not pool:
            bad.append(f"units.{u['id']} ({emo}/{tier}) has no persona to speak as")
    if bad:
        raise SystemExit("! workbook has broken references:\n  " + "\n  ".join(bad))
    print("  references OK — %d abilities, %d units, %d dialogue rows" %
          (len(ABILS), len(UNITS_ROWS), len(_dlg)))
_validate()

WB._sheets.sort(key=lambda s: ["README","emotions","abilities","loadouts","matchups","units","dialogue","prompts","moments",
  "layer_types","status_effects","synergies","rules","sounds","stations","metro_lines","travel_elements","items","armor","world_bands","city_status","checks"].index(s.title))
WB.save(book_path())
print("saved")
