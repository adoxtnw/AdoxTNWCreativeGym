#!/usr/bin/env python3
"""
Rebuild config/avui-config.xlsx from scratch.

    python3 tools/build_workbook.py

This is the generator for the WORKBOOK ITSELF — sheet layout, column sets,
LIVE/planned markers, validation formulas, and the seed content. Run it when
you want to add a sheet or a column. Day-to-day balancing happens in Google
Sheets instead; this would overwrite that.

Dialogue rows are imported from sources/neuro_metro_avui_enemy_dialogues.csv.
"""
import os
HERE = os.path.dirname(os.path.abspath(__file__))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WB = openpyxl.Workbook()
WB.remove(WB.active)

F = "Arial"
HDR_FILL   = PatternFill("solid", fgColor="1E1E26")
HDR_FONT   = Font(name=F, bold=True, color="F0ECE0", size=10)
LIVE_FILL  = PatternFill("solid", fgColor="D9EAD3")   # column read by the prototype today
PLAN_FILL  = PatternFill("solid", fgColor="FFF2CC")   # reserved for future design
NOTE_FONT  = Font(name=F, italic=True, color="666666", size=9)
BODY       = Font(name=F, size=10)
BOLD       = Font(name=F, bold=True, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def sheet(name, headers, live, rows, widths=None, notes=None):
    """headers: list of column names. live: set of names the game reads today."""
    ws = WB.create_sheet(name)
    if notes:
        ws["A1"] = notes; ws["A1"].font = NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers),6))
        hrow = 3
    else:
        hrow = 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORD
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        tag = ws.cell(row=hrow+1, column=c, value="LIVE" if h in live else "planned")
        tag.font = Font(name=F, size=8, bold=(h in live),
                        color="274E13" if h in live else "7F6000")
        tag.fill = LIVE_FILL if h in live else PLAN_FILL
        tag.border = BORD
        tag.alignment = Alignment(horizontal="left")
    for r, row in enumerate(rows, hrow+2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.font = BODY; cell.border = BORD
    for c, h in enumerate(headers, 1):
        w = (widths or {}).get(h, max(11, min(34, len(h) + 5)))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=hrow+2, column=2)
    return ws, hrow

# ───────────────────────────── README ─────────────────────────────
rd = WB.create_sheet("README")
lines = [
 ("NEURO-METRO: AVUI — GAME CONFIG", 16, True, "F0ECE0", "1E1E26"),
 ("", 10, False, None, None),
 ("This workbook is the source of truth for all game content. The prototype reads no", 10, False, None, None),
 ("content values of its own — every number, ability, unit and rule comes from here.", 10, False, None, None),
 ("", 10, False, None, None),
 ("HOW THE ROUND TRIP WORKS", 12, True, None, None),
 ("1.  Edit any sheet in Google Sheets.", 10, False, None, None),
 ("2.  File > Download > Comma-separated values, one CSV per sheet.", 10, False, None, None),
 ("3.  Put the CSVs in a folder and run:  python3 tools/build_data.py <folder>", 10, False, None, None),
 ("4.  That regenerates data.js. Reload the prototype. Nothing else to touch.", 10, False, None, None),
 ("", 10, False, None, None),
 ("COLUMN STATUS — the row under each header", 12, True, None, None),
 ("LIVE      (green)   the prototype reads this column today. Renaming it breaks the game.", 10, False, None, None),
 ("planned   (yellow)  reserved for design you have not built yet. Inert until wired up;", 10, False, None, None),
 ("                    safe to fill in early. Tell me when you want one activated.", 10, False, None, None),
 ("", 10, False, None, None),
 ("CONVENTIONS", 12, True, None, None),
 ("ids          UPPER_SNAKE. Referenced by other sheets, so renaming one means updating", 10, False, None, None),
 ("             every sheet that points at it. The 'checks' sheet catches broken references.", 10, False, None, None),
 ("lists        pipe-separated in a single cell:  ATK_ANGER|ATK_JOY|DEFEND", 10, False, None, None),
 ("booleans     1 or 0  (TRUE/FALSE also accepted)", 10, False, None, None),
 ("blank        means 'none' / 'typeless'. An ability with no emotion never matches a layer.", 10, False, None, None),
 ("*            wildcard, in the matchups sheet only. Matches any emotion.", 10, False, None, None),
 ("enabled      set to 0 to switch a row off without deleting it.", 10, False, None, None),
 ("", 10, False, None, None),
 ("ADDING THINGS — no code needed", 12, True, None, None),
 ("a new ability      add a row to 'abilities', then add its id to a unit's pool", 10, False, None, None),
 ("a new enemy        add a row to 'units'", 10, False, None, None),
 ("a new emotion      add a row to 'emotions' with a hex colour, then matchup rows for it", 10, False, None, None),
 ("a new synergy      add a row to 'matchups' — this is the main lever for interactions", 10, False, None, None),
 ("retune anything    'rules'", 10, False, None, None),
 ("", 10, False, None, None),
 ("SHEETS", 12, True, None, None),
 ("emotions         the six emotional types and their palette colours", 10, False, None, None),
 ("abilities        everything a unit can do", 10, False, None, None),
 ("matchups         attack emotion x layer emotion -> damage and charge. Your synergy table.", 10, False, None, None),
 ("units            player and enemies: stats, starting layers, ability pool", 10, False, None, None),
 ("layer_types      per-layer properties. Reserved — layers are plain emotions today.", 10, False, None, None),
 ("status_effects   buffs and debuffs. Reserved.", 10, False, None, None),
 ("synergies        line-level combos (sequences, counts, positions). Reserved.", 10, False, None, None),
 ("dialogue         what each enemy says, per persona and battle state", 10, False, None, None),
 ("rules            global tunables", 10, False, None, None),
 ("sounds           runtime-synthesised SFX. No audio files.", 10, False, None, None),
 ("checks           live validation. Every row should read OK.", 10, False, None, None),
]
for i,(txt,size,bold,fg,bg) in enumerate(lines, 1):
    c = rd.cell(row=i, column=1, value=txt)
    c.font = Font(name=F, size=size, bold=bold, color=fg or "000000")
    if bg: c.fill = PatternFill("solid", fgColor=bg)
rd.column_dimensions["A"].width = 100

# ───────────────────────────── emotions ─────────────────────────────
EMO = [
 # OFFICIAL palette, straight from Bars_Reference_001.svg. This list's ORDER is
 # also the gradient order (the reference puts them at even 20% stops).
 ("ANGER",   "Anger",   "anger",   "#e53859","#2a0810","Hot, direct, aggressive."),
 ("SURPRISE","Surprise","surprise","#724082","#170a1b","Sudden, disruptive, disorienting."),
 ("DISGUST", "Disgust", "disgust", "#56a36a","#0c1f13","Rejecting, corrosive, contemptuous."),
 ("JOY",     "Joy",     "joy",     "#fcc336","#2a1f06","Bright, volatile, energising."),
 ("SADNESS", "Sadness", "sadness", "#3d66c1","#0a1229","Heavy, draining, persistent."),
 ("FEAR",    "Fear",    "fear",    "#929fa5","#1a1e20","Cold, freezing, evasive."),
]
sheet("emotions",
  ["id","name","token","hex","bg_hex","short","description","enabled","notes"],
  {"id","name","token","hex","bg_hex"},
  [dict(id=i,name=n,token=t,hex=h,bg_hex=b,short=i[:3],description=d,enabled=1,notes="")
   for i,n,t,h,b,d in EMO],
  widths={"description":40,"notes":30},
  notes=("One row per emotional type. 'token' is the CSS palette variable name; 'hex' drives every colour "
         "in the UI for this emotion; 'bg_hex' tints the battle backdrop when an enemy of this type is fought."))

# ───────────────────────────── abilities ─────────────────────────────
ABIL_COLS = ["id","name","emotion","cost","kind","power","charge","hits_layer","icon",
             "target","reach","self_ms","self_ec","ec_push_target","ec_drain_target",
             "heal","shield_gain","pierce_shield","ignore_layer","repeat","cooldown",
             "uses","blurb","status_apply","status_chance","status_duration",
             "combo_tag","requires_prev_tag","synergy_group","wild_target",
             "rarity","unlock","enabled","suggested_cost","cost_delta","notes"]
ABIL_LIVE = {"id","name","emotion","cost","kind","power","charge","hits_layer","icon","self_ms","cooldown","uses","blurb","status_apply","status_duration"}
def ab(**k):
    row = {c:"" for c in ABIL_COLS}
    row.update(dict(target="ENEMY", reach="SINGLE", charge=0, repeat=1, cooldown=1,
                    uses=3, status_chance=0, status_duration=0,
                    pierce_shield=0, ignore_layer=0, enabled=1,
                    self_ms=0, self_ec=0, ec_push_target=0, ec_drain_target=0,
                    heal=0, shield_gain=0, wild_target="LOWEST_MS", rarity="COMMON"))
    row.update(k); return row
ABILS = [
 ab(id="ATK_ANGER", blurb='A straight {ANGER} strike. Costs {EC}, takes {MS} off the target and cracks its outermost {LAYER}.',  name="Anger",  emotion="ANGER",   cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="BOLT"),
 ab(id="ATK_SADNESS", blurb='A straight {SADNESS} strike. Costs {EC}, takes {MS} off the target and cracks its outermost {LAYER}.',name="Sadness",emotion="SADNESS", cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="DROP"),
 ab(id="ATK_JOY", blurb='A straight {JOY} strike. Costs {EC}, takes {MS} off the target and cracks its outermost {LAYER}.',    name="Joy",    emotion="JOY",     cost=20, kind="DAMAGE", power=35, hits_layer=1, icon="SPARK"),
 ab(id="DEFEND", blurb='Raises a guard. The next hit against you is *blocked* outright instead of costing {MS}.',     name="Defend", emotion="",        cost=10, kind="SHIELD", power=1,  hits_layer=0, icon="SHIELD",
    target="SELF", wild_target="SELF", shield_gain=1),
 ab(uses=2, id="HVY_ANGER", blurb='Heavy {ANGER}. Holds two *charge* segments first, so it lands late and hands the opponent room, but it hits for almost triple.',  name="Rage",   emotion="ANGER",   cost=45, kind="DAMAGE", power=90, charge=2, hits_layer=1, icon="BOLT"),
 ab(uses=2, id="HVY_SADNESS", blurb='Heavy {SADNESS}. Holds two *charge* segments first, so it lands late and hands the opponent room, but it hits for almost triple.',name="Grief",  emotion="SADNESS", cost=45, kind="DAMAGE", power=90, charge=2, hits_layer=1, icon="DROP"),
 ab(uses=2, id="HVY_JOY", blurb='Heavy {JOY}. Three *charge* segments and the hardest hit in the pool. Slow, loud, and expensive in every sense.',    name="Mania",  emotion="JOY",     cost=60, kind="DAMAGE", power=130,charge=3, hits_layer=1, icon="SPARK"),
 ab(id="RECHARGE", blurb='Buys {EC} with your own {MS}. Gains charge and lowers your ceiling at the same time — watch the {OVERLOAD} line.',   name="Recharge",emotion="",       cost=0,  kind="CHARGE", power=20, hits_layer=0, icon="CHARGE",
    target="SELF", wild_target="SELF", self_ms=5,
    notes="Gains 20 EC and costs 5 of your own MS — which also lowers your ceiling."),
 ab(id="GEN_DISGUST", blurb='Grows one extra {DISGUST} {LAYER} on yourself. *Grown layers never regrow* — once it breaks, it is gone.', name="Bile",   emotion="DISGUST", cost=15, kind="ADDLAYER", power=1, hits_layer=0, icon="ROT",
    target="SELF", wild_target="SELF", uses=2,
    notes="Grows one extra Disgust layer. Grown layers are TEMPORARY: once broken they never regrow."),
 ab(id="GEN_ANGER", blurb='Grows one extra {ANGER} {LAYER} on yourself. *Grown layers never regrow* — once it breaks, it is gone.',   name="Bristle",emotion="ANGER",   cost=15, kind="ADDLAYER", power=1, hits_layer=0, icon="ROT",
    target="SELF", wild_target="SELF", uses=2,
    notes="Grows one extra Anger layer. Grown layers are TEMPORARY: once broken they never regrow."),
 ab(id="ROT", blurb='{DISGUST} rot. For *2 turns* the target cannot regrow *2* of its broken {LAYERS}.',         name="Rot",    emotion="DISGUST", cost=25, kind="DEBUFF",   power=0, hits_layer=0, icon="ROT",
    uses=2, status_apply="NO_REGEN", status_duration=2,
    notes="For 2 turns the target cannot regrow 2 of its broken layers."),
 ab(id="INFLICT_SAD", blurb='{SADNESS} turned inward. For *2 turns* the target uses one of its own attacks *on itself* at the end of every round.', name="Self-Harm", emotion="SADNESS", cost=30, kind="DEBUFF", power=0, hits_layer=0, icon="DROP",
    uses=2, status_apply="SAD", status_duration=2,
    notes="Each round end the target turns one of its own attacks on itself. NOTE: distinct from "
          "SELF_HARM, which is the Overload-forced station."),
 ab(id="BLIND", blurb='{ANGER} blots out aim. For *2 turns* the target *misses half* of its attacks.',       name="Blinded by Hate", emotion="ANGER", cost=30, kind="DEBUFF", power=0, hits_layer=0, icon="EYE",
    uses=2, status_apply="BLINDED", status_duration=2,
    notes="ENEMY ability. The target misses half its attacks for 2 turns."),
 ab(cooldown=0, uses=0, id="SELF_HARM", blurb='Forced onto your line by {OVERLOAD}. Costs you {MS} and cannot be removed.',  name="Self Harm",emotion="",      cost=0,  kind="SELFHARM", power=25, hits_layer=0, icon="WARN",
    target="SELF", wild_target="SELF", rarity="OVERLOAD", enabled=1,
    notes="OVERLOAD ONLY. Forced into your line when Charge passes your ceiling. Cannot be moved or removed."),
 ab(cooldown=0, uses=0, id="FEED", blurb='Forced onto your line by {OVERLOAD}. *Heals your opponent* and cannot be removed.',       name="Feed",    emotion="",       cost=0,  kind="FEED",     power=30, hits_layer=0, icon="DROP",
    target="ENEMY", wild_target="AS_WRITTEN", rarity="OVERLOAD", enabled=1,
    notes="OVERLOAD ONLY. Heals your opponent. Cannot be moved or removed."),
]
ws_ab, hr = sheet("abilities", ABIL_COLS, ABIL_LIVE, ABILS,
  widths={"notes":34,"combo_tag":16,"requires_prev_tag":18,"synergy_group":16,"status_apply":16},
  notes=("One row per ability. kind: DAMAGE | SHIELD (add more kinds as you design them). "
         "power = damage for DAMAGE, shield charges for SHIELD. Leave 'emotion' blank for typeless — "
         "a typeless ability never matches a layer. 'suggested_cost' is a formula: a costing sanity check, not a rule."))
first = hr+2
for r in range(first, first+len(ABILS)):
    ws_ab.cell(row=r, column=ABIL_COLS.index("suggested_cost")+1).value = (
        f"=ROUND(G{r}*0.6 + N{r}*0.4 + P{r}*0.7 + Q{r}*8 + ABS(L{r})*0.3, 0)")
    ws_ab.cell(row=r, column=ABIL_COLS.index("cost_delta")+1).value = (
        f"=IFERROR(D{r}-AG{r},\"\")")
    for cc in ("suggested_cost","cost_delta"):
        ws_ab.cell(row=r, column=ABIL_COLS.index(cc)+1).font = BODY

# ───────────────────────────── matchups ─────────────────────────────
MU_COLS = ["attack_emotion","layer_emotion","priority","dmg_mult","ec_mult","label",
           "tag_class","sound","rotates_layer","self_ec","status_apply","enabled","notes"]
MU_LIVE = {"attack_emotion","layer_emotion","priority","dmg_mult","ec_mult","label","tag_class"}
def mu(a,l,p,d,e,lab,cls,snd,note="",rot="",sec=0,st=""):
    return dict(attack_emotion=a, layer_emotion=l, priority=p, dmg_mult=d, ec_mult=e,
                label=lab, tag_class=cls, sound=snd, rotates_layer=rot, self_ec=sec,
                status_apply=st, enabled=1, notes=note)
MUS = [
 mu("*","*",0,1,0,"OFF TYPE","off","hit","Fallback. Every pair you do not name lands here."),
 mu("*","NONE",5,2,0,"EXPOSED!","dmg","hit","No layers left: double damage, no charge."),
]
for e,_,_,_,_,_ in EMO:
    MUS.append(mu(e,e,10,0.5,1,"ABSORBED!","absorb","absorb","Same emotion: the layer drinks it."))
sheet("matchups", MU_COLS, MU_LIVE, MUS,
  widths={"notes":52,"label":16},
  notes=("THE SYNERGY TABLE — the main lever for designing interactions. One row per "
         "attack-emotion x layer-emotion pair. '*' is a wildcard; 'NONE' in layer_emotion means the "
         "target has no layers. When several rows match, the HIGHEST priority wins, so specific rows "
         "beat wildcards. Add a row like  JOY | SADNESS | 20 | 2.0 | 0 | CATHARSIS  and that pairing "
         "immediately behaves differently — no code change."))

# ───────────────────────────── units ─────────────────────────────
U_COLS = ["id","name","emotion","max_ms","start_ec_pct","layers","pool","line_dir",
          "line_cap","max_bonus_slots",
          "ai_profile","init","start_shield","max_layers_override","tags","enabled","notes"]
U_LIVE = {"id","name","emotion","max_ms","start_ec_pct","layers","pool","line_dir","line_cap","max_bonus_slots"}
UNITS_ROWS = [
 dict(id="player", name="You", emotion="", max_ms=400, start_ec_pct=0.40,
      layers="JOY|SADNESS", pool="ATK_ANGER|ATK_SADNESS|ATK_JOY|DEFEND|RECHARGE|HVY_ANGER|HVY_SADNESS|HVY_JOY|GEN_DISGUST|GEN_ANGER|ROT|INFLICT_SAD", line_dir=1,
      line_cap=3, max_bonus_slots=6,
      ai_profile="", init=10, start_shield=0, max_layers_override="", tags="PLAYER", enabled=1, notes=""),
 dict(id="enemy", name="The Commuter", emotion="ANGER", max_ms=250, start_ec_pct=0.40,
      layers="ANGER|ANGER|SADNESS", pool="ATK_ANGER|ATK_SADNESS|ATK_JOY|DEFEND|RECHARGE|HVY_ANGER|HVY_SADNESS|HVY_JOY|BLIND", line_dir=-1,
      line_cap=3, max_bonus_slots=6,
      ai_profile="GREEDY_MAX_DAMAGE", init=8, start_shield=0, max_layers_override="",
      tags="ENEMY", enabled=1, notes="AI reads the matchups sheet, so retuning it retunes the AI."),
]
sheet("units", U_COLS, U_LIVE, UNITS_ROWS,
  widths={"layers":26,"pool":40,"notes":42},
  notes=("One row per combatant. 'layers' is outermost-first and rotates as it takes hits. "
         "'pool' is which abilities this unit may use. Add enemies by adding rows."))

# ───────────────────────────── layer_types (reserved) ─────────────────────────────
LT_COLS = ["id","emotion","display_name","durability","on_break_effect","on_break_value",
           "passive_effect","passive_value","regrow_delay_ms","tags","enabled","notes"]
sheet("layer_types", LT_COLS, set(),
  [dict(id="EXAMPLE_BRITTLE_ANGER", emotion="ANGER", display_name="Brittle Anger", durability=1,
        on_break_effect="EC_TO_OWNER", on_break_value=10, passive_effect="DMG_TAKEN_MULT",
        passive_value=1.25, regrow_delay_ms=380, tags="EXAMPLE", enabled=0,
        notes="EXAMPLE ROW — shows the intended format. enabled=0, so it is ignored.")],
  widths={"notes":54,"on_break_effect":20,"passive_effect":20},
  notes=("RESERVED. Layers are plain emotions today — units.layers holds emotion ids. This sheet exists "
         "for when a layer needs properties of its own (taking several hits to break, doing something "
         "when it breaks, or granting a passive while it is outermost)."))

# ───────────────────────────── status_effects (reserved) ─────────────────────────────
SE_COLS = ["id","name","duration","icon","color","blurb","block_regen","miss_chance","self_hits",
           "dmg_taken_mult","dmg_dealt_mult","ec_gain_mult","ms_per_turn","ec_per_turn",
           "blocks_actions","stacking","max_stacks","enabled","notes"]
SE_LIVE = {"id","name","duration","icon","color","blurb","block_regen","miss_chance","self_hits","enabled"}
def se(**k):
    row = {c:"" for c in SE_COLS}
    row.update(dict(duration=2, block_regen=0, miss_chance=0, self_hits=0,
                    dmg_taken_mult=1, dmg_dealt_mult=1, ec_gain_mult=1,
                    ms_per_turn=0, ec_per_turn=0, blocks_actions=0,
                    stacking="REFRESH", max_stacks=1, enabled=1))
    row.update(k); return row
SE_ROWS = [
 se(id="NO_REGEN", name="Rotting", duration=2, icon="ROT", color="#56a36a", block_regen=2,
    blurb="*2* of your broken {LAYERS} are being held down. They stay broken until this wears off.",
    notes="Holds this many broken layers down: they stay broken while it lasts."),
 se(id="SAD",      name="Sad",     duration=2, icon="DROP", color="#3d66c1", self_hits=1,
    blurb="At the end of every round one of your own attacks is turned *on yourself*.",
    notes="Each round end the victim turns one of its OWN attacks on itself."),
 se(id="BLINDED",  name="Blinded", duration=2, icon="EYE", color="#e53859", miss_chance=0.5,
    blurb="{ANGER} has blotted out your aim. *Half* of your attacks miss outright.",
    notes="The victim fluffs this fraction of its attacks."),
]
sheet("status_effects", SE_COLS, SE_LIVE, SE_ROWS,
  widths={"notes":52},
  notes=("One row per status. Abilities apply these by id through status_apply / "
         "status_duration. Add a status by adding a row and teaching Kinds what reads it."))


# ───────────────────────────── synergies (reserved) ─────────────────────────────
SY_COLS = ["id","name","trigger","subject","operator","value","window","effect",
           "effect_value","applies_to","priority","enabled","notes"]
sheet("synergies", SY_COLS, set(),
  [dict(id="EXAMPLE_TRIPLE", name="Monochrome Line", trigger="LINE_CONTAINS", subject="EMOTION_COUNT",
        operator=">=", value=3, window="LINE", effect="DMG_MULT", effect_value=1.5,
        applies_to="ALL_STATIONS", priority=10, enabled=0,
        notes="EXAMPLE ROW — three stations of one emotion in a line hit 50% harder. enabled=0."),
   dict(id="EXAMPLE_CHAIN", name="Grief After Rage", trigger="PREV_STATION", subject="EMOTION",
        operator="==", value="ANGER", window="STATION", effect="EC_MULT", effect_value=2,
        applies_to="THIS_STATION", priority=20, enabled=0,
        notes="EXAMPLE ROW — a station firing straight after an Anger station gives double charge. enabled=0.")],
  widths={"notes":56},
  notes=("RESERVED. Line-level combos: things that depend on what else is in the line, or on what "
         "fired just before. trigger/subject/operator/value form the condition; effect/effect_value "
         "form the result. Tell me the ones you want and I will wire the vocabulary you use here."))

# ───────────────────────────── rules ─────────────────────────────
RULES = [
 ("lineCap",3,"FALLBACK line size, used only if a unit row leaves line_cap blank."),
 ("maxLayers",6,"Layer slots per unit."),
 ("restEc",15,"Charge gained by departing with an empty line."),
 ("ecBasis","POWER","POWER = charge figured from the ability's base power. DEALT = from damage actually dealt."),
 ("noLayerDmgMult",1,"Superseded by the matchups sheet; kept as a fallback."),
 ("noLayerEcMult",0,"Superseded by the matchups sheet; kept as a fallback."),
 ("shieldRotatesLayer",0,"Whether a blocked attack still rotates the layer queue."),
 ("lowMsWarnPct",0.18,"Below this fraction of max MS the ceiling chevron pulses red."),
 ("clashSeconds",4,"Length of the death sequence."),
 ("layerRegenAtTurnEnd",1,"Broken layers stay gone until the round ends, then regrow at the back of the queue."),
 ("chargeStepMs",260,"How long each charge segment holds before the next."),
 ("flyMs",420,"How long a station takes to fly across and strike the target."),
 ("typeMs",32,"Delay between letters of dialogue."),
 ("dialogueHoldMs",3000,"How long a finished line stays on screen before it fades."),
 ("lowHpTalkPct",0.2,"Fraction of max MS below which the winning / losing lines fire."),
 ("overloadSlotPer",0.25,"Every this-much overflow (as a fraction of max MS) forces one more corrupted slot."),
 ("aiVarietyChance",0.22,"Chance the AI takes a random affordable attack instead of its best one."),
 ("aiDebuffChance",0.5,"Chance the AI spends a slot on a debuff the target is not already suffering."),
 ("aiChargeBias",0.55,"Chance the AI prefers an ability with charge segments when it can afford one."),
 ("shuffleLayersEachRound",1,"Re-order every unit's layer queue at the start of each round."),
 ("chargeGrantsSlots",1,"Each charge segment in a line grants the OPPONENT one temporary slot next turn."),
 ("maxBonusSlots",6,"FALLBACK cap on temporary slots, used only if a unit row leaves max_bonus_slots blank."),
 ("abilPageSize",4,"Abilities shown per page in the Emotions panel."),
("slotArriveMs",260,"Gap between one temporary slot flying in and the next."),
 ("unsheathMs",340,"How long one interface element takes to draw itself into place."),
 ("unsheathGapMs",110,"Gap before the next element starts — smaller than unsheathMs, so they overlap."),
 ("longPressMs",420,"How long an ability must be held before its tooltip opens."),
 ("swipeMinPx",36,"How far a pointer must travel across the panel to count as a page swipe."),
 ("tagGrowth",0.85,"How much an accumulating tag grows once it holds a full max-MS worth of change."),
 ("tagBreathSlowMs",1700,"Breathing period of an accumulating tag holding almost nothing."),
 ("tagBreathFastMs",380,"Breathing period of a tag holding a full max-MS worth — it flashes far harder."),
 ("gaugeW",134,"Gauge canvas width in logical pixels. Lower = chunkier pixels, coarser value steps."),
 ("gaugeH",22,"Gauge canvas height in logical pixels — tall enough for the end caps."),
 ("barCoreH",12,"Height of the bar's interior, before the wave, in logical pixels."),
 ("barWaveAmp",1.4,"Wave amplitude on the bar's silhouette (reference: 9.8 of a 1026-long bar)."),
 ("barWaveHalf",5.6,"Half-period of the silhouette wave, i.e. one arc (reference: 43.3)."),
 ("barWaveSpeed",0.10,"How fast the silhouette wave travels."),
 ("ecBandFrac",0.82,"Charge fill height as a fraction of the bar interior (reference: 76 of 92.74)."),
 ("ecWaveAmp",0.7,"Amplitude of the charge fill's own wave — long and shallow, per the reference."),
 ("ecWaveHalf",19,"Half-period of the charge fill's wave. Much longer than the silhouette's."),
 ("ecWaveSpeed",0.05,"How fast the charge wave travels."),
 ("ecInnerFrac",0.5,"Inner charge bar thickness, as a fraction of the outer one. Same shape, dodged over it."),
 ("ecInnerAmt",0.4,"How hard the inner charge bar dodges. 1 blows out to white; lower keeps the hue."),
 ("arcSpacing",3.1,"Spacing of the background arc texture (reference: 23.83)."),
 ("arcRadius",12.6,"Radius of the background arcs (reference: 96.95)."),
 ("arcAlpha",0.11,"How strongly the arcs show. They only read against the unlit stretch."),
 ("capW",5,"End-cap width in logical pixels."),
 ("capR",2,"End-cap corner rounding in logical pixels."),
 ("ballFlyMs",330,"How long a ball of light takes to arc from its indicator into the bar."),
 ("settleStepMs",70,"Pause between one hit landing on a bar and the next setting off."),
 ("barTweenMs",1000,"How long a bar takes to travel to its new value once a hit lands."),
 ("overloadHoldMs",2000,"How long the OVERLOAD tag and its slow motion last."),
 ("ecScrollSpeed",0.0022,"How fast the charge gradient drifts along the bar."),
 ("themeOpening","audio/theme-opening.wav","Theme part 1: plays once."),
 ("themeLoop","audio/theme-loop.wav","Theme part 2: loops for ever, scheduled to start the instant part 1 ends."),
 ("musicFadeMs",900,"How quickly the theme fades out when the battle ends."),
 ("musicVolume",0.30,"Theme level. Lower this if the sound effects are getting buried."),
 ("sfxVolume",1.05,"Sound-effect level, mixed against musicVolume."),
 ("layerEase",0.16,"How fast layers slide to their new slot. Higher = snappier."),
 ("layerWaveDelay",0.85,"Phase offset per slot, so breathing travels outward in waves."),
 ("layerInnerShrink",0.85,"Curve of the radius falloff toward the centre."),
 ("layerFill",0.42,"Ring thickness as a fraction of the gap to the next slot."),
 ("layerOuterThick",2,"Thickness multiplier for the outermost (active) layer."),
 ("layerInnerThick",0.9,"Thickness multiplier for every layer behind it."),
 ("layerFlashMs",240,"How long a struck layer flashes white before vanishing."),
 ("layerGapMs",260,"Beat between the layer vanishing and it regrowing at the back."),
 ("layerRegrowMs",380,"How long the regrow beat holds before resolution continues."),
 ("enemyRingBaseR",28,"Enemy: radius of the outermost ring, in logical canvas pixels."),
 ("enemyRingSpacing",4.7,"Enemy: radius step between one layer and the next in. baseR/spacing sets how many fit."),
 ("enemyRingBreathe",1.6,"Enemy ring breathing amplitude."),
 ("playerRingBaseR",300,"Player: radius of the outermost ring. Huge on purpose — a big circle reads as a nearly flat band."),
 ("playerRingSpacing",5,"Player: radius step between layers. Small, so all six sit in a thin arc despite the huge radius."),
 ("playerRingBreathe",2.2,"Player ring breathing amplitude."),
 ("playerCanvasW",120,"Player ring canvas width in logical pixels (wide and short)."),
 ("playerCanvasH",32,"Player ring canvas height in logical pixels."),
 ("playerRingCy",300,"Player: circle centre Y, pushed far below the strip so only the flat top of the arc shows."),
 ("crusherSteps",12,"Audio bit-crusher quantisation. Low = Atari harsh, high = SNES-clean."),
]
sheet("rules", ["key","value","description"], {"key","value"},
  [dict(key=k, value=v, description=d) for k,v,d in RULES],
  widths={"key":24,"value":14,"description":78},
  notes="Global tunables. Keys are read by name — renaming one breaks whatever reads it.")

# ───────────────────────────── sounds ─────────────────────────────
SND = [
 ("tap","square",980,1460,80,0.46,0.30,"Adding an ability to the line."),
 ("place","square",620,1760,130,0.40,0.35,"The station lands on your line and flashes white."),
 ("sheatheE","sawtooth",1500,220,150,0.30,0.22,"Enemy UI element drawn into place. Lower."),
 ("sheatheP","sawtooth",2100,340,140,0.28,0.22,"Player UI element drawn into place. Higher."),
 ("slot","triangle",300,1500,150,0.30,0.45,"A temporary slot flies in from the opponent."),
 ("remove","square",520,220,90,0.34,0.24,"Removing a station."),
 ("depart","sawtooth",240,70,420,0.26,0.45,"The line departs and charge is spent."),
 ("travel","square",150,150,38,0.10,0.15,"Line scrolling to the next station."),
 ("arrive","square",880,880,45,0.14,0.30,"Station reaches the head and lights up."),
 ("hit","noise",700,700,190,0.34,0.40,"Off-type hit."),
 ("absorb","square",300,1200,260,0.24,0.50,"Same-type hit: the layer absorbs it."),
 ("block","noise",1600,1600,90,0.30,0.35,"Shield blocks, or shield goes up."),
 ("breaklayer","sawtooth",760,80,340,0.28,0.55,"A layer flashes and breaks."),
 ("regrow","triangle",120,520,300,0.16,0.40,"The broken layer regrows at the back."),
 ("clash","noise",240,240,1500,0.40,0.65,"Death: the chevrons meet."),
 ("speak","square",620,540,42,0.09,0.12,"One letter of dialogue appearing."),
 ("win","square",440,1320,620,0.24,0.55,"Victory."),
 ("lose","sawtooth",320,50,950,0.28,0.60,"Defeat."),
]
sheet("sounds", ["id","wave","f0","f1","dur","gain","echo","description"],
  {"id","wave","f0","f1","dur","gain","echo"},
  [dict(id=i,wave=w,f0=a,f1=b,dur=d,gain=g,echo=e,description=n) for i,w,a,b,d,g,e,n in SND],
  widths={"description":42},
  notes=("Every sound is synthesised at runtime — there are no audio files. "
         "wave: square | sawtooth | triangle | sine | noise. f0/f1 sweep in Hz over 'dur' ms. "
         "'echo' is how much of the sound feeds the shared delay line. Everything runs through a "
         "bit-crusher, which is where the crunch comes from."))

# ───────────────────────────── dialogue ─────────────────────────────
import csv as _csv
_dlg_src = os.path.join(HERE, "..", "sources", "neuro_metro_avui_enemy_dialogues.csv")
_dlg = []
if os.path.exists(_dlg_src):
    for r in _csv.DictReader(open(_dlg_src, encoding="utf-8-sig")):
        _dlg.append(dict(emotion=r["Emotion"].strip().upper(),
                         persona=r["Persona"].strip(),
                         state=r["State"].strip().upper(),
                         line=r["Dialogue Line"].strip(),
                         enabled=1, notes=""))
# ───────────────────────────── moments ─────────────────────────────
# The title screen names Barcelona's current moment. `day` is a weekday or "*",
# and from_hour/to_hour are BARCELONA local hours, to_hour exclusive; a band that
# wraps past midnight (22 -> 5) is written as from=22 to=5. Higher priority wins,
# so a Friday-night line beats the everyday one; ties are picked at random.
MOM_COLS = ["id","day","from_hour","to_hour","phrase","priority","enabled","notes"]
MOM_LIVE = {"id","day","from_hour","to_hour","phrase","priority","enabled"}
def mo(i, day, a, b, phrase, pri=0):
    return dict(id=i, day=day, from_hour=a, to_hour=b, phrase=phrase,
                priority=pri, enabled=1, notes="")
MOMENTS = [
 # ---- everyday, by band ----
 mo("dawn_1","*",5,7,"the first light is still deciding"),
 mo("dawn_2","*",5,7,"nobody has spoken yet today"),
 mo("morn_1","*",7,10,"early morning"),
 mo("morn_2","*",7,10,"the early carriages are full of unfinished sleep"),
 mo("morn_3","*",7,10,"everyone is already late"),
 mo("mid_1","*",10,13,"middle of the day"),
 mo("mid_2","*",10,13,"the light is flat and honest"),
 mo("noon_1","*",13,16,"all afternoon ahead"),
 mo("noon_2","*",13,16,"the long slow part of the day"),
 mo("late_1","*",16,19,"the afternoon is running out"),
 mo("late_2","*",16,19,"everyone is going somewhere they did not choose"),
 mo("dusk_1","*",19,21,"somber twilight"),
 mo("dusk_2","*",19,21,"the hour the day admits what it was"),
 mo("night_1","*",21,24,"the night is getting loud"),
 mo("night_2","*",21,24,"the last trains are filling up"),
 mo("small_1","*",0,5,"the madrugada, and still moving"),
 mo("small_2","*",0,5,"too late to be anything but honest"),
 # ---- weekday specifics ----
 mo("mon_morn","MON",5,10,"monday, and the week already weighs something",2),
 mo("mon_dusk","MON",19,24,"the longest monday of the month",2),
 mo("wed_mid","WED",10,16,"the week is stuck exactly in the middle",2),
 mo("thu_late","THU",16,21,"almost, but not yet",2),
 mo("fri_dusk","FRI",19,21,"a night to finally be free",2),
 mo("fri_night","FRI",21,24,"a night to finally be free",2),
 mo("fri_small","FRI",0,5,"friday refusing to end",2),
 mo("sat_mid","SAT",10,16,"a saturday with nothing owed to anyone",2),
 mo("sat_night","SAT",21,24,"the city is pretending it never has to work again",2),
 mo("sun_mid","SUN",10,16,"a sunday that stretches too far",2),
 mo("sun_dusk","SUN",19,24,"sunday dusk, and monday already breathing on it",2),
]
sheet("moments", MOM_COLS, MOM_LIVE, MOMENTS,
  widths={"phrase":54,"notes":30},
  notes=("Title-screen mood line: \"Barcelona, <weekday>, <phrase>.\" day is MON..SUN "
         "or * for any. Hours are Barcelona local, to_hour exclusive; a band may wrap "
         "past midnight. Highest priority wins, ties are random."))

sheet("dialogue", ["emotion","persona","state","line","enabled","notes"],
  {"emotion","persona","state","line"}, _dlg,
  widths={"line":92,"persona":20,"notes":20},
  notes=("What each enemy says. states: INTRO (battle start) - WINNING (player drops below 20% MS) - "
         "LOSING (this enemy drops below 20% MS) - DEFEAT (this enemy dies). A battle picks one persona "
         "at random from the rows matching the enemy's emotion, so the same enemy type speaks differently "
         "each run. NOTE: this sheet uses Anger/Disgust/Sadness/Fear/Joy/Surprise, which is NOT the same "
         "six as the emotions sheet - only Anger, Sadness and Joy line up. See the README."))

# ───────────────────────────── checks ─────────────────────────────
ck = WB.create_sheet("checks")
ck["A1"] = "LIVE VALIDATION — every row should read OK. Anything else means a broken reference."
ck["A1"].font = NOTE_FONT
hdr = ["check","expected","actual","status"]
for c,h in enumerate(hdr,1):
    x = ck.cell(row=3, column=c, value=h); x.font = HDR_FONT; x.fill = HDR_FILL; x.border = BORD
CHECKS = [
 ("Abilities with an emotion not in 'emotions'", 0,
  '=SUMPRODUCT((abilities!C5:C200<>"")*(COUNTIF(emotions!A5:A200,abilities!C5:C200)=0))'),
 ("Abilities missing an id", 0, '=SUMPRODUCT((abilities!B5:B200<>"")*(abilities!A5:A200=""))'),
 ("Units with a blank pool", 0, '=SUMPRODUCT((units!A5:A200<>"")*(units!G5:G200=""))'),
 ("Matchup rows with no label", 0, '=SUMPRODUCT((matchups!A5:A200<>"")*(matchups!F5:F200=""))'),
 ("Matchup attack emotions unknown (blank/*/NONE allowed)", 0,
  '=SUMPRODUCT((matchups!A5:A200<>"")*(matchups!A5:A200<>"*")*(matchups!A5:A200<>"NONE")*(COUNTIF(emotions!A5:A200,matchups!A5:A200)=0))'),
 ("Matchup layer emotions unknown (blank/*/NONE allowed)", 0,
  '=SUMPRODUCT((matchups!B5:B200<>"")*(matchups!B5:B200<>"*")*(matchups!B5:B200<>"NONE")*(COUNTIF(emotions!A5:A200,matchups!B5:B200)=0))'),
 ("Duplicate ability ids", 0,
  '=SUMPRODUCT((abilities!A5:A200<>"")*(COUNTIF(abilities!A5:A200,abilities!A5:A200&"")>1))'),
 ("Duplicate emotion ids", 0,
  '=SUMPRODUCT((emotions!A5:A200<>"")*(COUNTIF(emotions!A5:A200,emotions!A5:A200&"")>1))'),
 ("Enabled emotions", 6, '=SUMPRODUCT((emotions!A5:A200<>"")*(emotions!G5:G200=1))'),
 ("Enabled abilities", 10, '=SUMPRODUCT((abilities!A5:A200<>"")*(abilities!AF5:AF200=1))'),
]
for r,(label,exp,formula) in enumerate(CHECKS, 4):
    ck.cell(row=r, column=1, value=label).font = BODY
    ck.cell(row=r, column=2, value=exp).font = BODY
    ck.cell(row=r, column=3, value=formula).font = BODY
    ck.cell(row=r, column=4, value=f'=IF(C{r}=B{r},"OK","CHECK")').font = BOLD
    for c in range(1,5): ck.cell(row=r, column=c).border = BORD
ck.column_dimensions["A"].width = 56
for col,w in (("B",12),("C",12),("D",12)): ck.column_dimensions[col].width = w

WB._sheets.sort(key=lambda s: ["README","emotions","abilities","matchups","units","dialogue","moments",
  "layer_types","status_effects","synergies","rules","sounds","checks"].index(s.title))
WB.save(os.path.join(HERE, "..", "config", "avui-config.xlsx"))
print("saved")
